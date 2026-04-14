"""
Puls — Ukentlig timerapportering
Start: uvicorn puls.app:app --reload --port 8502
Eller: python puls/app.py
"""
from fastapi import FastAPI, Request, Query, Cookie
from typing import Optional
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from jinja2 import Environment, FileSystemLoader, select_autoescape
from markupsafe import Markup
from pathlib import Path
import json
import sqlite3
from contextlib import contextmanager
from datetime import datetime, date, timedelta
import uvicorn
import openpyxl
import os
import secrets
import calendar

app = FastAPI(title="Puls")
BASE = Path(__file__).parent
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "puls-admin")
EXPORT_API_KEY = os.environ.get("EXPORT_API_KEY", "")

# Jinja2 direkte (omgår Starlette-wrapper som har bug i Python 3.14)
jinja_env = Environment(
    loader=FileSystemLoader(BASE / "templates"),
    autoescape=select_autoescape(["html"]),
    cache_size=0,
)
jinja_env.filters["tojson"] = lambda v: Markup(json.dumps(v, ensure_ascii=False))

def render(template_name: str, **ctx) -> HTMLResponse:
    html = jinja_env.get_template(template_name).render(**ctx)
    return HTMLResponse(html)

# På Azure App Service bruker vi /home/data (persistent på tvers av restarts)
# Lokalt bruker vi puls/data/
DATA_DIR = Path("/home/data") if os.environ.get("WEBSITE_SITE_NAME") else BASE / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_FIL   = DATA_DIR / "puls.db"
FAKTA_FIL = DATA_DIR / "fakta_puls.xlsx"

# ── Database ──────────────────────────────────────────────────────────────────

@contextmanager
def db():
    con = sqlite3.connect(DB_FIL, timeout=15)
    con.row_factory = sqlite3.Row
    try:
        yield con
        con.commit()
    finally:
        con.close()

def init_db():
    with db() as con:
        con.executescript("""
            CREATE TABLE IF NOT EXISTS brukere (
                token TEXT PRIMARY KEY,
                navn  TEXT NOT NULL,
                epost TEXT NOT NULL,
                lønn  INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS investeringer (
                navn       TEXT PRIMARY KEY,
                rekkefølge INTEGER NOT NULL DEFAULT 0,
                kategori   TEXT NOT NULL DEFAULT 'Annet'
            );
            CREATE TABLE IF NOT EXISTS svar (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                token     TEXT NOT NULL,
                navn      TEXT NOT NULL,
                epost     TEXT NOT NULL,
                uke       INTEGER NOT NULL,
                år        INTEGER NOT NULL,
                fravar    INTEGER NOT NULL DEFAULT 0,
                timer     TEXT NOT NULL DEFAULT '{}',
                total     REAL NOT NULL DEFAULT 0,
                tidspunkt TEXT NOT NULL,
                UNIQUE(token, uke, år)
            );
            CREATE TABLE IF NOT EXISTS trivsel_utsendelser (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                måned      INTEGER NOT NULL,
                år         INTEGER NOT NULL,
                opprettet  TEXT NOT NULL,
                åpen_dager INTEGER NOT NULL DEFAULT 10,
                stengt     INTEGER NOT NULL DEFAULT 0,
                UNIQUE(måned, år)
            );
            CREATE TABLE IF NOT EXISTS trivsel_tokens (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                survey_token  TEXT UNIQUE NOT NULL,
                utsendelse_id INTEGER NOT NULL,
                bruker_token  TEXT NOT NULL,
                brukt         INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS trivsel_svar (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                utsendelse_id INTEGER NOT NULL,
                trivsel       INTEGER NOT NULL,
                samarbeid     INTEGER NOT NULL,
                innsendt      TEXT NOT NULL
            );
        """)
        # Migrering: legg til kategori-kolonne hvis den ikke finnes
        cols = [r[1] for r in con.execute("PRAGMA table_info(investeringer)").fetchall()]
        if "kategori" not in cols:
            con.execute("ALTER TABLE investeringer ADD COLUMN kategori TEXT NOT NULL DEFAULT 'Annet'")
        # Migrering: legg til lønn og team på brukere
        cols_b = [r[1] for r in con.execute("PRAGMA table_info(brukere)").fetchall()]
        if "lønn" not in cols_b:
            con.execute("ALTER TABLE brukere ADD COLUMN lønn INTEGER NOT NULL DEFAULT 0")
        if "team" not in cols_b:
            con.execute("ALTER TABLE brukere ADD COLUMN team TEXT NOT NULL DEFAULT 'investering'")
        if "aktiv" not in cols_b:
            con.execute("ALTER TABLE brukere ADD COLUMN aktiv INTEGER NOT NULL DEFAULT 1")
        # Migrering: trivsel_svar fra gammelt skjema (runde_id/spm1/spm2) til nytt
        cols_ts = [r[1] for r in con.execute("PRAGMA table_info(trivsel_svar)").fetchall()]
        if cols_ts and "utsendelse_id" not in cols_ts:
            con.execute("DROP TABLE trivsel_svar")
            con.execute("""
                CREATE TABLE trivsel_svar (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    utsendelse_id INTEGER NOT NULL,
                    trivsel       INTEGER NOT NULL,
                    samarbeid     INTEGER NOT NULL,
                    innsendt      TEXT NOT NULL
                )
            """)

init_db()

# ── DB-hjelpefunksjoner ───────────────────────────────────────────────────────

KATEGORIER = ["Laks", "Sjømat", "Investeringer", "Kapital", "Annet"]

DEFAULT_INVESTERINGER = [
    {"navn": "SalMar",            "kategori": "Laks"},
    {"navn": "Sinkaberg-Hansen",  "kategori": "Laks"},
    {"navn": "Arctic Fish",       "kategori": "Laks"},
    {"navn": "Kingfish Company",  "kategori": "Laks"},
    {"navn": "LaxValoris",        "kategori": "Laks"},
    {"navn": "Scale",             "kategori": "Sjømat"},
    {"navn": "Pelagia",           "kategori": "Sjømat"},
    {"navn": "Insula",            "kategori": "Sjømat"},
    {"navn": "BEWi",              "kategori": "Investeringer"},
    {"navn": "Salvesen & Thams",  "kategori": "Investeringer"},
    {"navn": "Kvarv",             "kategori": "Annet"},
    {"navn": "Kverva-møter",      "kategori": "Annet"},
    {"navn": "Styremøter",        "kategori": "Annet"},
    {"navn": "Admin / Annet",     "kategori": "Annet"},
]

def les_investeringer() -> list[dict]:
    """Returnerer liste av dicts: [{navn, kategori}, ...]"""
    with db() as con:
        rader = con.execute("SELECT navn, kategori FROM investeringer ORDER BY rekkefølge").fetchall()
    if not rader:
        return DEFAULT_INVESTERINGER
    return [{"navn": r["navn"], "kategori": r["kategori"]} for r in rader]

def les_inv_navn() -> list[str]:
    """Kun navneliste — for bakoverkompatibel bruk."""
    return [i["navn"] for i in les_investeringer()]

def lagre_investeringer(liste: list[dict]):
    """liste = [{navn, kategori}, ...]"""
    with db() as con:
        con.execute("DELETE FROM investeringer")
        for i, item in enumerate(liste):
            con.execute(
                "INSERT OR REPLACE INTO investeringer (navn, rekkefølge, kategori) VALUES (?,?,?)",
                (item["navn"], i, item.get("kategori", "Annet"))
            )

def finn_bruker(token: str):
    with db() as con:
        r = con.execute("SELECT navn, epost FROM brukere WHERE token=?", (token,)).fetchone()
    return dict(r) if r else None

def hent_alle_brukere() -> dict:
    with db() as con:
        rader = con.execute("SELECT token, navn, epost, lønn, team, aktiv FROM brukere").fetchall()
    return {r["token"]: {"navn": r["navn"], "epost": r["epost"], "lønn": r["lønn"], "team": r["team"], "aktiv": r["aktiv"]} for r in rader}

def sett_aktiv_bruker(token: str, aktiv: int):
    with db() as con:
        con.execute("UPDATE brukere SET aktiv=? WHERE token=?", (aktiv, token))

def lagre_bruker(token: str, navn: str, epost: str):
    with db() as con:
        con.execute("INSERT OR REPLACE INTO brukere (token, navn, epost) VALUES (?,?,?)", (token, navn, epost))

def sett_lønn_bruker(token: str, lønn: int):
    with db() as con:
        con.execute("UPDATE brukere SET lønn=? WHERE token=?", (lønn, token))

def sett_team_bruker(token: str, team: str):
    with db() as con:
        con.execute("UPDATE brukere SET team=? WHERE token=?", (team, token))

def fjern_bruker(token: str) -> str:
    with db() as con:
        r = con.execute("SELECT navn FROM brukere WHERE token=?", (token,)).fetchone()
        navn = r["navn"] if r else token
        con.execute("DELETE FROM brukere WHERE token=?", (token,))
    return navn

def _rad_til_svar(r) -> dict:
    return {
        "token": r["token"],
        "navn": r["navn"],
        "epost": r["epost"],
        "uke": r["uke"],
        "år": r["år"],
        "fravær": bool(r["fravar"]),
        "timer": json.loads(r["timer"]),
        "total": r["total"],
        "tidspunkt": r["tidspunkt"],
    }

def hent_alle_svar() -> list:
    with db() as con:
        rader = con.execute("SELECT * FROM svar ORDER BY tidspunkt").fetchall()
    return [_rad_til_svar(r) for r in rader]

def upsert_svar(token, navn, epost, uke, år, fravar, timer, total, tidspunkt):
    with db() as con:
        con.execute("""
            INSERT INTO svar (token, navn, epost, uke, år, fravar, timer, total, tidspunkt)
            VALUES (?,?,?,?,?,?,?,?,?)
            ON CONFLICT(token, uke, år) DO UPDATE SET
                navn=excluded.navn, epost=excluded.epost,
                fravar=excluded.fravar, timer=excluded.timer,
                total=excluded.total, tidspunkt=excluded.tidspunkt
        """, (token, navn, epost, uke, år, int(fravar), json.dumps(timer, ensure_ascii=False), total, tidspunkt))

# ── Hjelpefunksjoner (uendret logikk) ────────────────────────────────────────

FAKTA_KOLONNER = ["Navn", "Epost", "Uke", "År", "Dato innsending", "Investering", "Timer"]

def skriv_fakta_puls(navn, epost, uke, år, tidspunkt, timer: dict):
    if FAKTA_FIL.exists():
        wb = openpyxl.load_workbook(FAKTA_FIL)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Puls"
        ws.append(FAKTA_KOLONNER)

    dato_str = tidspunkt[:10]
    rader_å_beholde = [ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [c.value for c in row]
        if not (vals[0] == navn and vals[2] == uke and vals[3] == år):
            rader_å_beholde.append(row)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Puls"
    ws2.append(FAKTA_KOLONNER)
    for row in rader_å_beholde[1:]:
        ws2.append([c.value for c in row])
    for inv, t in timer.items():
        ws2.append([navn, epost, uke, år, dato_str, inv, t])
    wb2.save(FAKTA_FIL)

def get_uke_år():
    iso = date.today().isocalendar()
    return iso[1], iso[0]

def forrige_uke_svar(token: str, uke: int, år: int) -> dict:
    fu, få = (52, år - 1) if uke == 1 else (uke - 1, år)
    with db() as con:
        r = con.execute("SELECT timer FROM svar WHERE token=? AND uke=? AND år=?", (token, fu, få)).fetchone()
    return json.loads(r["timer"]) if r else {}

def har_svart(token: str, uke: int, år: int) -> bool:
    with db() as con:
        r = con.execute("SELECT 1 FROM svar WHERE token=? AND uke=? AND år=?", (token, uke, år)).fetchone()
    return r is not None

def historikk_bruker(token: str, år: int) -> list:
    with db() as con:
        rader = con.execute("SELECT * FROM svar WHERE token=? AND år=? ORDER BY uke", (token, år)).fetchall()
    return [_rad_til_svar(r) for r in rader]

def siste_svar(token: str, uke: int, år: int) -> dict | None:
    with db() as con:
        r = con.execute("""
            SELECT * FROM svar WHERE token=? AND NOT (uke=? AND år=?)
            ORDER BY år DESC, uke DESC LIMIT 1
        """, (token, uke, år)).fetchone()
    return _rad_til_svar(r) if r else None

def fredag_kl_12(uke: int, år: int) -> datetime:
    jan4 = date(år, 1, 4)
    mandag = jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=uke - 1)
    fredag = mandag + timedelta(days=4)
    return datetime(fredag.year, fredag.month, fredag.day, 12, 0, 0)

def fmt_delta(minutter: float) -> str:
    m = int(minutter)
    if m < 60:
        return f"{m} min"
    return f"{m // 60}t {m % 60:02d}min"

def ranker_uke(uke: int, år: int) -> list:
    t0 = fredag_kl_12(uke, år)
    with db() as con:
        rader = con.execute(
            "SELECT * FROM svar WHERE uke=? AND år=? AND fravar=0", (uke, år)
        ).fetchall()
    resultat = []
    for r in rader:
        s = _rad_til_svar(r)
        delta = max(0, (datetime.fromisoformat(s["tidspunkt"]) - t0).total_seconds() / 60)
        resultat.append({"navn": s["navn"].split()[0], "delta_min": delta, "delta_fmt": fmt_delta(delta), "total": s.get("total", 0)})
    return sorted(resultat, key=lambda x: x["delta_min"])

def måneds_ranking(måned: int, år: int) -> list:
    with db() as con:
        rader = con.execute("SELECT * FROM svar WHERE år=? AND fravar=0", (år,)).fetchall()
    per_person: dict = {}
    for r in rader:
        s = _rad_til_svar(r)
        if datetime.fromisoformat(s["tidspunkt"]).month != måned:
            continue
        t0 = fredag_kl_12(s["uke"], s["år"])
        delta = max(0, (datetime.fromisoformat(s["tidspunkt"]) - t0).total_seconds() / 60)
        navn = s["navn"].split()[0]
        per_person.setdefault(navn, []).append(delta)
    resultat = [{"navn": n, "snitt_min": sum(v) / len(v), "snitt_fmt": fmt_delta(sum(v) / len(v)), "antall": len(v)} for n, v in per_person.items()]
    return sorted(resultat, key=lambda x: x["snitt_min"])[:5]

def all_time_toppliste() -> list:
    with db() as con:
        uker = con.execute("SELECT DISTINCT uke, år FROM svar").fetchall()
    poeng: dict = {}
    for row in uker:
        for i, r in enumerate(ranker_uke(row["uke"], row["år"])):
            p = poeng.setdefault(r["navn"], {"poeng": 0, "nr1": 0, "antall": 0})
            p["antall"] += 1
            p["nr1"] += (i == 0)
            p["poeng"] += max(0, 5 - i)
    return sorted([{"navn": n, **v} for n, v in poeng.items()], key=lambda x: -x["poeng"])[:8]

def hall_of_shame_liste(nå_uke: int, nå_år: int) -> list:
    brukere = {t: b for t, b in hent_alle_brukere().items() if b["aktiv"]}
    resultat = []
    for token, b in brukere.items():
        with db() as con:
            rapporterte = {(r["uke"], r["år"]) for r in con.execute(
                "SELECT uke, år FROM svar WHERE token=?", (token,)
            ).fetchall()}
        mangler_n = sum(1 for u in range(1, nå_uke) if (u, nå_år) not in rapporterte)
        if mangler_n > 0:
            resultat.append({"navn": b["navn"].split()[0], "mangler": mangler_n})
    return sorted(resultat, key=lambda x: -x["mangler"])[:5]

def personlig_stats(token: str, nå_uke: int, nå_år: int) -> dict:
    with db() as con:
        rader = con.execute(
            "SELECT * FROM svar WHERE token=? AND år=? AND fravar=0", (token, nå_år)
        ).fetchall()
    mine = [_rad_til_svar(r) for r in rader]
    if not mine:
        return {}
    total_timer = sum(s.get("total", 0) for s in mine)
    antall_uker = len(mine)
    inv_sum: dict = {}
    for s in mine:
        for inv, t in s.get("timer", {}).items():
            inv_sum[inv] = inv_sum.get(inv, 0) + t
    favoritt = max(inv_sum, key=lambda k: inv_sum[k]) if inv_sum else "–"
    streak = 0
    uker_svart = {s["uke"] for s in mine}
    for u in range(nå_uke - 1, 0, -1):
        if u in uker_svart:
            streak += 1
        else:
            break
    return {
        "antall_uker": antall_uker,
        "total_timer": total_timer,
        "snitt": round(total_timer / antall_uker, 1) if antall_uker else 0,
        "favoritt": favoritt,
        "streak": streak,
    }

def manglende_uker(token: str, nå_uke: int, nå_år: int) -> list:
    with db() as con:
        rapporterte = {(r["uke"], r["år"]) for r in con.execute(
            "SELECT uke, år FROM svar WHERE token=?", (token,)
        ).fetchall()}
    return [(u, nå_år) for u in range(1, nå_uke) if (u, nå_år) not in rapporterte]

# ── Kostnadsfordeling ────────────────────────────────────────────────────────

MÅNEDS_NAVN = ["Januar","Februar","Mars","April","Mai","Juni",
               "Juli","August","September","Oktober","November","Desember"]

def finn_uker_for_måned(måned: int, år: int) -> list:
    _, days = calendar.monthrange(år, måned)
    uker = set()
    for dag in range(1, days + 1):
        iso = date(år, måned, dag).isocalendar()
        uker.add((iso[1], iso[0]))
    return list(uker)

def hent_alle_timer_for_uker(token: str, uker: list) -> dict:
    """Returnerer alle timer inkl. Annet-kategorier."""
    timer_inv: dict = {}
    for (uke, å) in uker:
        with db() as con:
            r = con.execute("SELECT * FROM svar WHERE token=? AND uke=? AND år=?", (token, uke, å)).fetchone()
        if r:
            s = _rad_til_svar(r)
            if not s["fravær"]:
                for inv, t in s["timer"].items():
                    if t > 0:
                        timer_inv[inv] = timer_inv.get(inv, 0) + t
    return timer_inv

def hent_timer_for_uker(token: str, uker: list, inkl_navn: set) -> dict:
    timer_inv: dict = {}
    for (uke, å) in uker:
        with db() as con:
            r = con.execute("SELECT * FROM svar WHERE token=? AND uke=? AND år=?", (token, uke, å)).fetchone()
        if r:
            s = _rad_til_svar(r)
            if not s["fravær"]:
                for inv, t in s["timer"].items():
                    if inv in inkl_navn and t > 0:
                        timer_inv[inv] = timer_inv.get(inv, 0) + t
    return timer_inv

def hent_ytd_snitt(token: str, aktuell_måned: int, år: int, inkl_navn: set) -> dict:
    måneder_data = []
    for m in range(1, aktuell_måned):
        uker = finn_uker_for_måned(m, år)
        t = hent_timer_for_uker(token, uker, inkl_navn)
        if t:
            måneder_data.append(t)
    if not måneder_data:
        return {}
    snitt: dict = {}
    for md in måneder_data:
        for inv, t in md.items():
            snitt[inv] = snitt.get(inv, 0) + t
    return {inv: t / len(måneder_data) for inv, t in snitt.items()}

def beregn_fordeling(total_kostnad: float, måned: int, år: int) -> dict:
    uker = finn_uker_for_måned(måned, år)
    inkl_inv = [i for i in les_investeringer() if i["kategori"] != "Annet"]
    inkl_navn = {i["navn"] for i in inkl_inv}
    inv_order = {i["navn"]: idx for idx, i in enumerate(inkl_inv)}

    with db() as con:
        brukere_rader = con.execute("SELECT token, navn, lønn, team FROM brukere ORDER BY navn").fetchall()

    # lønn=0 → telles som 1 (lik vekt)
    total_lønn = sum(max(b["lønn"] or 0, 1) for b in brukere_rader)

    personer = []
    for b in brukere_rader:
        lønn = max(b["lønn"] or 0, 1)
        andel = lønn / total_lønn
        kostnad_person = total_kostnad * andel
        team = b["team"] or "investering"

        if team == "investering":
            # Investeringsteam: kun inkluderte investeringer, YTD-fallback
            timer = hent_timer_for_uker(b["token"], uker, inkl_navn)
            brukt_ytd = False
            if sum(timer.values()) == 0:
                timer = hent_ytd_snitt(b["token"], måned, år, inkl_navn)
                brukt_ytd = bool(timer)
            total_timer = sum(timer.values())
            inv_kostnad_person: dict = {}
            if total_timer > 0:
                for inv, t in timer.items():
                    inv_kostnad_person[inv] = kostnad_person * (t / total_timer)
            personer.append({
                "token": b["token"], "navn": b["navn"], "lønn": b["lønn"] or 0,
                "team": team, "andel_prosent": round(andel * 100, 2),
                "kostnad_person": round(kostnad_person),
                "timer": timer, "total_timer": round(total_timer, 1),
                "annet_timer": 0, "annet_kostnad": 0,
                "inv_kostnad": inv_kostnad_person,
                "brukt_ytd": brukt_ytd, "brukt_team_nøkkel": False,
                "ingen_timer": total_timer == 0,
            })

        else:  # støtte
            # Hent ALLE timer inkl. Annet for å finne riktig proporsjon
            alle_timer = hent_alle_timer_for_uker(b["token"], uker)
            alle_total = sum(alle_timer.values())
            inkl_timer = {k: v for k, v in alle_timer.items() if k in inkl_navn}
            inkl_total = sum(inkl_timer.values())
            annet_total = alle_total - inkl_total

            inv_kostnad_person = {}
            annet_kostnad = 0.0

            if alle_total > 0:
                # Steg 1: direkte allokering fra investeringstimer
                for inv, t in inkl_timer.items():
                    inv_kostnad_person[inv] = kostnad_person * (t / alle_total)
                # Steg 2: resten (Annet-timer) → team-nøkkel i neste pass
                annet_kostnad = kostnad_person * (annet_total / alle_total)
            else:
                # Ingen timer i det hele tatt → alt til team-nøkkel
                annet_kostnad = kostnad_person

            personer.append({
                "token": b["token"], "navn": b["navn"], "lønn": b["lønn"] or 0,
                "team": team, "andel_prosent": round(andel * 100, 2),
                "kostnad_person": round(kostnad_person),
                "timer": inkl_timer, "total_timer": round(inkl_total, 1),
                "annet_timer": round(annet_total, 1), "annet_kostnad": annet_kostnad,
                "inv_kostnad": inv_kostnad_person,
                "brukt_ytd": False, "brukt_team_nøkkel": False,
                "ingen_timer": alle_total == 0,
            })

    # Bygg team-nøkkel fra alle direkte allokerte investeringskostnader
    inv_kr_direkte: dict = {}
    totalt_kr_direkte = 0.0
    for p in personer:
        for inv, kr in p["inv_kostnad"].items():
            inv_kr_direkte[inv] = inv_kr_direkte.get(inv, 0.0) + kr
            totalt_kr_direkte += kr

    team_nøkkel = {inv: kr / totalt_kr_direkte for inv, kr in inv_kr_direkte.items()} if totalt_kr_direkte > 0 else {}

    # Støtte: fordel annet_kostnad via team-nøkkel
    for p in personer:
        if p["team"] == "støtte" and p["annet_kostnad"] > 0 and team_nøkkel:
            for inv, nøkkel in team_nøkkel.items():
                p["inv_kostnad"][inv] = p["inv_kostnad"].get(inv, 0.0) + p["annet_kostnad"] * nøkkel
            p["brukt_team_nøkkel"] = True
            p["ingen_timer"] = False

    # Summer kostnad per investering på tvers av alle personer
    inv_kostnad_total: dict = {}
    for p in personer:
        for inv, kr in p["inv_kostnad"].items():
            inv_kostnad_total[inv] = inv_kostnad_total.get(inv, 0.0) + kr

    resultat = []
    for inv_navn in sorted(inv_kostnad_total.keys(), key=lambda n: inv_order.get(n, 999)):
        kr = inv_kostnad_total[inv_navn]
        kat = next((i["kategori"] for i in inkl_inv if i["navn"] == inv_navn), "")
        resultat.append({
            "investering": inv_navn,
            "kategori": kat,
            "prosent": round(kr / total_kostnad * 100, 2) if total_kostnad else 0,
            "kostnad": round(kr),
        })

    totalt_fordelt = sum(r["kostnad"] for r in resultat)

    return {
        "resultat": resultat,
        "total_lønn": total_lønn,
        "totalt_fordelt": totalt_fordelt,
        "total_kostnad": total_kostnad,
        "personer": personer,
        "måned": måned,
        "år": år,
        "måned_navn": MÅNEDS_NAVN[måned - 1],
    }

# ── Ruter ────────────────────────────────────────────────────────────────────

@app.get("/puls/{token}", response_class=HTMLResponse)
async def vis_skjema(request: Request, token: str,
                     uke: Optional[int] = Query(None),
                     år: Optional[int] = Query(None)):
    bruker = finn_bruker(token)
    if not bruker:
        return HTMLResponse("<h1 style='font-family:sans-serif;padding:40px'>Ugyldig eller utløpt lenke.</h1>", status_code=404)
    nå_uke, nå_år = get_uke_år()
    uke = uke if uke is not None else nå_uke
    år  = år  if år  is not None else nå_år
    allerede_svart = har_svart(token, uke, år)
    forrige = forrige_uke_svar(token, uke, år)
    hist    = historikk_bruker(token, år)
    mangler = manglende_uker(token, nå_uke, nå_år)
    mangler = [(u, å) for u, å in mangler if not (u == uke and å == år)]
    siste   = siste_svar(token, uke, år)
    return render("form.html",
        bruker=bruker, token=token, uke=uke, år=år,
        investeringer=les_investeringer(), forrige=forrige,
        allerede_svart=allerede_svart, historikk=hist,
        mangler=mangler, siste=siste,
    )

@app.post("/puls/{token}", response_class=HTMLResponse)
async def send_inn(request: Request, token: str):
    bruker = finn_bruker(token)
    if not bruker:
        return HTMLResponse("<h1>Ugyldig lenke</h1>", status_code=404)
    form = await request.form()
    nå_uke, nå_år = get_uke_år()
    try:
        uke = int(form.get("_uke", nå_uke))
        år  = int(form.get("_år",  nå_år))
    except (ValueError, TypeError):
        uke, år = nå_uke, nå_år
    investeringer = les_inv_navn()
    timer = {}
    total = 0
    for inv in investeringer:
        v = min(40, max(0, int(form.get(f"t_{inv.replace(' ','_').replace('/','_')}", 0) or 0)))
        timer[inv] = v
        total += v
    fravar = form.get("_fravar") == "1"

    if fravar:
        upsert_svar(token, bruker["navn"], bruker["epost"], uke, år, True, {}, 0, datetime.now().isoformat())
        return RedirectResponse(f"/puls/{token}/takk?uke={uke}&år={år}&fravar=1", status_code=303)

    if total > 40:
        total = 40
    upsert_svar(token, bruker["navn"], bruker["epost"], uke, år, False, timer, total, datetime.now().isoformat())
    return RedirectResponse(f"/puls/{token}/takk?uke={uke}&år={år}", status_code=303)

@app.get("/puls/{token}/takk", response_class=HTMLResponse)
async def takk(request: Request, token: str,
               uke: Optional[int] = Query(None),
               år: Optional[int] = Query(None),
               fravar: Optional[int] = Query(None)):
    bruker = finn_bruker(token)
    if not bruker:
        return HTMLResponse("<h1>Ugyldig lenke</h1>", status_code=404)
    nå_uke, nå_år = get_uke_år()
    uke = uke if uke is not None else nå_uke
    år  = år  if år  is not None else nå_år
    hist = historikk_bruker(token, år)
    siste = next((s for s in reversed(hist) if s["uke"] == uke), None)
    mangler = manglende_uker(token, nå_uke, nå_år)
    return render("takk.html",
        bruker=bruker, token=token, uke=uke, år=år,
        historikk=hist, siste=siste, mangler=mangler,
        fravar=bool(fravar),
    )

@app.get("/puls/{token}/stats", response_class=HTMLResponse)
async def stats(request: Request, token: str):
    bruker = finn_bruker(token)
    if not bruker:
        return HTMLResponse("<h1 style='font-family:sans-serif;padding:40px'>Ugyldig lenke.</h1>", status_code=404)
    nå_uke, nå_år = get_uke_år()
    nå_måned = date.today().month
    return render("stats.html",
        bruker=bruker, token=token, uke=nå_uke, år=nå_år,
        denne_uken=ranker_uke(nå_uke, nå_år),
        måneds=måneds_ranking(nå_måned, nå_år),
        nå_måned_navn=date(nå_år, nå_måned, 1).strftime("%B %Y").capitalize(),
        alltime=all_time_toppliste(),
        shame=hall_of_shame_liste(nå_uke, nå_år),
        mine=personlig_stats(token, nå_uke, nå_år),
        historikk=historikk_bruker(token, nå_år),
    )

# ── Admin ────────────────────────────────────────────────────────────────────

ADMIN_COOKIE = "puls_admin"

def er_innlogget(request: Request) -> bool:
    token = request.cookies.get(ADMIN_COOKIE, "")
    return secrets.compare_digest(token, ADMIN_PASSWORD)

@app.get("/admin", response_class=HTMLResponse)
async def admin_get(request: Request):
    innlogget = er_innlogget(request)
    return render("admin.html",
        innlogget=innlogget,
        feil=False,
        melding=request.query_params.get("melding", ""),
        brukere=hent_alle_brukere() if innlogget else {},
        investeringer=les_investeringer() if innlogget else [],
    )

@app.post("/admin/login")
async def admin_login(request: Request):
    form = await request.form()
    passord = form.get("passord", "")
    if secrets.compare_digest(passord, ADMIN_PASSWORD):
        response = RedirectResponse("/admin", status_code=303)
        response.set_cookie(ADMIN_COOKIE, ADMIN_PASSWORD, httponly=True, samesite="lax")
        return response
    return render("admin.html", innlogget=False, feil=True, melding="", brukere={}, investeringer=[])

@app.get("/admin/logout")
async def admin_logout():
    response = RedirectResponse("/admin", status_code=303)
    response.delete_cookie(ADMIN_COOKIE)
    return response

@app.post("/admin/brukere/legg-til")
async def admin_legg_til_bruker(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    token = form.get("token", "").strip().lower()
    navn  = form.get("navn", "").strip()
    epost = form.get("epost", "").strip()
    if token and navn and epost:
        lagre_bruker(token, navn, epost)
        # Auto-opprett trivsel-token for alle åpne utsendelser inneværende år
        nå = datetime.now()
        with db() as con:
            åpne = con.execute(
                "SELECT id FROM trivsel_utsendelser WHERE år=? AND stengt=0", (nå.year,)
            ).fetchall()
            for u in åpne:
                eks = con.execute(
                    "SELECT 1 FROM trivsel_tokens WHERE utsendelse_id=? AND bruker_token=?",
                    (u["id"], token)
                ).fetchone()
                if not eks:
                    con.execute(
                        "INSERT INTO trivsel_tokens (survey_token, utsendelse_id, bruker_token) VALUES (?,?,?)",
                        (secrets.token_urlsafe(24), u["id"], token)
                    )
    return RedirectResponse(f"/admin?melding={navn}+lagt+til", status_code=303)

@app.post("/admin/brukere/fjern")
async def admin_fjern_bruker(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    token = form.get("token", "").strip()
    navn = fjern_bruker(token)
    return RedirectResponse(f"/admin?melding={navn}+fjernet", status_code=303)


@app.get("/admin/migrering-2026")
async def admin_migrering(request: Request):
    if not er_innlogget(request):
        return JSONResponse({"feil": "ikke innlogget"})
    import base64, json as _json
    DATA_B64 = "eyJicnVrZXJlIjogW1siYW11bmQiLCAiQW11bmQgR2FsYWVuIiwgImFtdW5kLmdhbGFlbkBrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJib3JnZSIsICJCw7hyZ2UgS2x1bmdlcmJvIiwgImJrQGt2ZXJ2YS5ubyIsICJpbnZlc3RlcmluZyJdLCBbImVpcmlrIiwgIkVpcmlrIFZhYm8iLCAiZWlyaWsudmFib0BrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJlcmlrIiwgIkVyaWsgRWhyZW5wb2hsIFNhbmQiLCAiZXJpay5laHJlbnBvaGwuc2FuZEBrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJmcm9kZSIsICJGcm9kZSBTYW5kbWFyayIsICJmcm9kZS5zYW5kbWFya0BrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJtYWdudXMiLCAiTWFnbnVzIER5YnZhZCIsICJtZEBrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJtYXRzLmciLCAiTWF0cyBHYWJyaWVsc2VuIiwgIm1hdHMuZ2FicmllbHNlbkBrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJtYXRzLm0iLCAiTWF0cyBNYWx2aWciLCAibWF0cy5tYWx2aWdAa3ZlcnZhLm5vIiwgImludmVzdGVyaW5nIl0sIFsibW9ydGVuIiwgIk1vcnRlbiBNasO4ZW4iLCAibWJtQGt2ZXJ2YS5ubyIsICJpbnZlc3RlcmluZyJdLCBbIm5pa29sYWkiLCAiTmlrb2xhaSBKZW5zZW4iLCAibmlrb2xhaS5qZW5zZW5Aa3ZlcnZhLm5vIiwgImludmVzdGVyaW5nIl0sIFsib2xhdiIsICJPbGF2IEhvbHN0IER5cm5lcyIsICJvbGF2LmhvbHN0LmR5cm5lc0BrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJwZXJuaWxsZSIsICJQZXJuaWxsZSBTa2Fyc3RlaW4iLCAicGVybmlsbGVAa3ZlcnZhLm5vIiwgImludmVzdGVyaW5nIl0sIFsicGlhIiwgIlBpYSBIYW1tZXIiLCAicGlhLmhhbW1lckBrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJyYWduYSIsICJSYWduYSBGYWxrYW5nZXIiLCAicmFnbmEuZmFsa2FuZ2VyQGt2ZXJ2YS5ubyIsICJpbnZlc3RlcmluZyJdLCBbInNpbWVuIiwgIlNpbWVuIE5pbHNlbiIsICJzb25Aa3ZlcnZhLm5vIiwgImludmVzdGVyaW5nIl0sIFsic3lubmUiLCAiU3lubmUgSMOlcnN0YWQiLCAic3lubmVAa3ZlcnZhLm5vIiwgImludmVzdGVyaW5nIl0sIFsidG9yb2xhdiIsICJUb3IgT2xhdiBBbmRlcnNlbiIsICJ0b2JhQGt2ZXJ2YS5ubyIsICJpbnZlc3RlcmluZyJdLCBbInRvcmdlaXIiLCAiVG9yZ2VpciBTdmFlIiwgInRvcmdlaXIuc3ZhZUBrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXSwgWyJ0b3JzdGVpbiIsICJUb3JzdGVpbiBaYWhsIiwgInR6ZUBrdmVydmEubm8iLCAiaW52ZXN0ZXJpbmciXV0sICJzdmFyIjogW1sibWFnbnVzIiwgIk1hZ251cyBEeWJ2YWQiLCAibWRAa3ZlcnZhLm5vIiwgMSwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAxLCBcIk51dHJpbWFyXCI6IDEsIFwiUmFwcG9ydGVyaW5nXCI6IDF9IiwgMy4wLCAiMjAyNi0wMS0wMlQwOToyMjo1MyJdLCBbInRvcnN0ZWluIiwgIlRvcnN0ZWluIFphaGwiLCAidHplQGt2ZXJ2YS5ubyIsIDEsIDIwMjYsIDAsICJ7XCJSYXBwb3J0ZXJpbmdcIjogNX0iLCA1LjAsICIyMDI2LTAxLTAyVDEyOjMzOjM1Il0sIFsiZXJpayIsICJFcmlrIEVocmVucG9obCBTYW5kIiwgImVyaWsuZWhyZW5wb2hsLnNhbmRAa3ZlcnZhLm5vIiwgMSwgMjAyNiwgMCwgIntcIk5vcmRpc2tlIGFrc2plclwiOiAyNCwgXCJTcGFyZWJhbmtlblwiOiA4LCBcIkdsb2JhbGUgYWtzamVyXCI6IDh9IiwgNDAuMCwgIjIwMjYtMDEtMDVUMDI6MjU6MDEiXSwgWyJhbXVuZCIsICJBbXVuZCBHYWxhZW4iLCAiYW11bmQuZ2FsYWVuQGt2ZXJ2YS5ubyIsIDEsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMS0wNVQwMjozMTo0MSJdLCBbInN5bm5lIiwgIlN5bm5lIEjDpXJzdGFkIiwgInN5bm5lQGt2ZXJ2YS5ubyIsIDIsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMS0wOVQwNjowMjo0NSJdLCBbIm1vcnRlbiIsICJNb3J0ZW4gTWrDuGVuIiwgIm1ibUBrdmVydmEubm8iLCAyLCAyMDI2LCAwLCAie1wiU2NhbGVcIjogMzMsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiA3fSIsIDQwLjAsICIyMDI2LTAxLTA5VDA2OjA2OjEzIl0sIFsicGlhIiwgIlBpYSBIYW1tZXIiLCAicGlhLmhhbW1lckBrdmVydmEubm8iLCAyLCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogNDB9IiwgNDAuMCwgIjIwMjYtMDEtMDlUMDY6MTI6NTQiXSwgWyJvbGF2IiwgIk9sYXYgSG9sc3QgRHlybmVzIiwgIm9sYXYuaG9sc3QuZHlybmVzQGt2ZXJ2YS5ubyIsIDIsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogMTgsIFwiQ29uY2hpbGlhXCI6IDgsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiAxMH0iLCAzNi4wLCAiMjAyNi0wMS0wOVQwNjoyMzoxOCJdLCBbIm5pa29sYWkiLCAiTmlrb2xhaSBKZW5zZW4iLCAibmlrb2xhaS5qZW5zZW5Aa3ZlcnZhLm5vIiwgMiwgMjAyNiwgMCwgIntcIlNjYWxlXCI6IDMzLCBcIkluc3VsYVwiOiA2LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMX0iLCA0MC4wLCAiMjAyNi0wMS0wOVQwNjoyNDoxNyJdLCBbImZyb2RlIiwgIkZyb2RlIFNhbmRtYXJrIiwgImZyb2RlLnNhbmRtYXJrQGt2ZXJ2YS5ubyIsIDIsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDMsIFwiTnV0cmltYXJcIjogNCwgXCJCZW5jaG1hcmtcIjogNywgXCJMYXhWYWxvcmlzXCI6IDEyLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogNCwgXCJLdmVydmEtbcO4dGVyXCI6IDEwfSIsIDQwLjAsICIyMDI2LTAxLTA5VDA3OjE2OjA0Il0sIFsicmFnbmEiLCAiUmFnbmEgRmFsa2FuZ2VyIiwgInJhZ25hLmZhbGthbmdlckBrdmVydmEubm8iLCAyLCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAxMCwgXCJOdXRyaW1hclwiOiAxNSwgXCJCZW5jaG1hcmtcIjogMTAsIFwiSVZDXCI6IDIsIFwiUmFwcG9ydGVyaW5nXCI6IDN9IiwgNDAuMCwgIjIwMjYtMDEtMDlUMDc6NTU6MTciXSwgWyJhbXVuZCIsICJBbXVuZCBHYWxhZW4iLCAiYW11bmQuZ2FsYWVuQGt2ZXJ2YS5ubyIsIDIsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMS0wOVQwODoxNjo0OSJdLCBbImVpcmlrIiwgIkVpcmlrIFZhYm8iLCAiZWlyaWsudmFib0BrdmVydmEubm8iLCAyLCAyMDI2LCAwLCAie1wiU2FsTWFyXCI6IDIwLCBcIlNpbmthYmVyZy1IYW5zZW5cIjogMjAsIFwiS3ZlcnZhLW3DuHRlclwiOiA4fSIsIDQ4LjAsICIyMDI2LTAxLTA5VDA5OjMyOjA3Il0sIFsidG9yc3RlaW4iLCAiVG9yc3RlaW4gWmFobCIsICJ0emVAa3ZlcnZhLm5vIiwgMiwgMjAyNiwgMCwgIntcIkxheFZhbG9yaXNcIjogMjAsIFwiUmFwcG9ydGVyaW5nXCI6IDIwfSIsIDQwLjAsICIyMDI2LTAxLTA5VDEwOjM1OjUwIl0sIFsidG9yb2xhdiIsICJUb3IgT2xhdiBBbmRlcnNlbiIsICJ0b2JhQGt2ZXJ2YS5ubyIsIDIsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiAzM30iLCAzMy4wLCAiMjAyNi0wMS0xMVQwMzowMzo1MCJdLCBbImVyaWsiLCAiRXJpayBFaHJlbnBvaGwgU2FuZCIsICJlcmlrLmVocmVucG9obC5zYW5kQGt2ZXJ2YS5ubyIsIDIsIDIwMjYsIDAsICJ7XCJOb3JkaXNrZSBha3NqZXJcIjogMjQsIFwiU3BhcmViYW5rZW5cIjogNSwgXCJHbG9iYWxlIGFrc2plclwiOiA1LCBcIlJhcHBvcnRlcmluZ1wiOiA1LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMX0iLCA0MC4wLCAiMjAyNi0wMS0xMlQwMzo1NTowOCJdLCBbIm1hZ251cyIsICJNYWdudXMgRHlidmFkIiwgIm1kQGt2ZXJ2YS5ubyIsIDIsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMTAsIFwiTnV0cmltYXJcIjogMTUsIFwiS3ZlcnZhLW3DuHRlclwiOiAxNX0iLCA0MC4wLCAiMjAyNi0wMS0xMlQxMzozNjo1NyJdLCBbInBlcm5pbGxlIiwgIlBlcm5pbGxlIFNrYXJzdGVpbiIsICJwZXJuaWxsZUBrdmVydmEubm8iLCAyLCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDMyLCBcIkdsb2JhbGUgYWtzamVyXCI6IDMsIFwiS3ZlcnZhLW3DuHRlclwiOiA1fSIsIDQwLjAsICIyMDI2LTAxLTE2VDEyOjI4OjA5Il0sIFsidG9yc3RlaW4iLCAiVG9yc3RlaW4gWmFobCIsICJ0emVAa3ZlcnZhLm5vIiwgMywgMjAyNiwgMCwgIntcIkxheFZhbG9yaXNcIjogMTAsIFwiUmVnbnNrYXBcIjogMjAsIFwiUmFwcG9ydGVyaW5nXCI6IDEwfSIsIDQwLjAsICIyMDI2LTAxLTE2VDA2OjAzOjU4Il0sIFsic3lubmUiLCAiU3lubmUgSMOlcnN0YWQiLCAic3lubmVAa3ZlcnZhLm5vIiwgMywgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAxLTE2VDA2OjA2OjIxIl0sIFsidG9yZ2VpciIsICJUb3JnZWlyIFN2YWUiLCAidG9yZ2Vpci5zdmFlQGt2ZXJ2YS5ubyIsIDMsIDIwMjYsIDAsICJ7XCJTY2FsZVwiOiA4LCBcIkluc3VsYVwiOiAyNSwgXCJJVkNcIjogNiwgXCJLdmVydmEtbcO4dGVyXCI6IDZ9IiwgNDUuMCwgIjIwMjYtMDEtMTZUMDY6MDg6MzQiXSwgWyJlcmlrIiwgIkVyaWsgRWhyZW5wb2hsIFNhbmQiLCAiZXJpay5laHJlbnBvaGwuc2FuZEBrdmVydmEubm8iLCAzLCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDI4LCBcIlNwYXJlYmFua2VuXCI6IDYsIFwiR2xvYmFsZSBha3NqZXJcIjogNCwgXCJLdmVydmEtbcO4dGVyXCI6IDJ9IiwgNDAuMCwgIjIwMjYtMDEtMTZUMDY6MDg6NDAiXSwgWyJtYWdudXMiLCAiTWFnbnVzIER5YnZhZCIsICJtZEBrdmVydmEubm8iLCAzLCAyMDI2LCAwLCAie1wiU2FsTWFyXCI6IDEwLCBcIk51dHJpbWFyXCI6IDEwLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMjB9IiwgNDAuMCwgIjIwMjYtMDEtMTZUMDY6MTM6MTkiXSwgWyJhbXVuZCIsICJBbXVuZCBHYWxhZW4iLCAiYW11bmQuZ2FsYWVuQGt2ZXJ2YS5ubyIsIDMsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMS0xNlQwNjoxNDoyNyJdLCBbImZyb2RlIiwgIkZyb2RlIFNhbmRtYXJrIiwgImZyb2RlLnNhbmRtYXJrQGt2ZXJ2YS5ubyIsIDMsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDIsIFwiTnV0cmltYXJcIjogNSwgXCJCZW5jaG1hcmtcIjogMTAsIFwiTGF4VmFsb3Jpc1wiOiAxMCwgXCJOeWUgY2FzZSBTasO4bWF0XCI6IDMsIFwiS3ZlcnZhLW3DuHRlclwiOiAxMH0iLCA0MC4wLCAiMjAyNi0wMS0xNlQwNjoyNToyOSJdLCBbInBpYSIsICJQaWEgSGFtbWVyIiwgInBpYS5oYW1tZXJAa3ZlcnZhLm5vIiwgMywgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAxLTE2VDA2OjM2OjU5Il0sIFsib2xhdiIsICJPbGF2IEhvbHN0IER5cm5lcyIsICJvbGF2LmhvbHN0LmR5cm5lc0BrdmVydmEubm8iLCAzLCAyMDI2LCAwLCAie1wiSW5zdWxhXCI6IDMyLCBcIkNvbmNoaWxpYVwiOiA1LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogM30iLCA0MC4wLCAiMjAyNi0wMS0xNlQwNjo0NTowNSJdLCBbIm1vcnRlbiIsICJNb3J0ZW4gTWrDuGVuIiwgIm1ibUBrdmVydmEubm8iLCAzLCAyMDI2LCAwLCAie1wiU2NhbGVcIjogMzcsIFwiSVZDXCI6IDN9IiwgNDAuMCwgIjIwMjYtMDEtMTZUMDc6NDA6NTAiXSwgWyJ0b3JvbGF2IiwgIlRvciBPbGF2IEFuZGVyc2VuIiwgInRvYmFAa3ZlcnZhLm5vIiwgMywgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDMyfSIsIDMyLjAsICIyMDI2LTAxLTE2VDEwOjU4OjI2Il0sIFsicGVybmlsbGUiLCAiUGVybmlsbGUgU2thcnN0ZWluIiwgInBlcm5pbGxlQGt2ZXJ2YS5ubyIsIDMsIDIwMjYsIDAsICJ7XCJOb3JkaXNrZSBha3NqZXJcIjogMjUsIFwiR2xvYmFsZSBha3NqZXJcIjogMiwgXCJLdmVydmEtbcO4dGVyXCI6IDEwLCBcIkt2YXJ2XCI6IDN9IiwgNDAuMCwgIjIwMjYtMDEtMTZUMTI6MjY6MjUiXSwgWyJlaXJpayIsICJFaXJpayBWYWJvIiwgImVpcmlrLnZhYm9Aa3ZlcnZhLm5vIiwgMywgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAzMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDUsIFwiUmFwcG9ydGVyaW5nXCI6IDUsIFwiS3ZlcnZhLW3DuHRlclwiOiA1fSIsIDQ1LjAsICIyMDI2LTAxLTE5VDAyOjI1OjM3Il0sIFsibmlrb2xhaSIsICJOaWtvbGFpIEplbnNlbiIsICJuaWtvbGFpLmplbnNlbkBrdmVydmEubm8iLCAzLCAyMDI2LCAwLCAie1wiU2NhbGVcIjogMTAsIFwiSW5zdWxhXCI6IDE0LCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMTMsIFwiS3ZlcnZhLW3DuHRlclwiOiAzfSIsIDQwLjAsICIyMDI2LTAxLTE5VDAyOjI3OjIxIl0sIFsibW9ydGVuIiwgIk1vcnRlbiBNasO4ZW4iLCAibWJtQGt2ZXJ2YS5ubyIsIDQsIDIwMjYsIDAsICJ7XCJTY2FsZVwiOiAzNSwgXCJJVkNcIjogNX0iLCA0MC4wLCAiMjAyNi0wMS0yM1QwNjowMTo1NSJdLCBbImFtdW5kIiwgIkFtdW5kIEdhbGFlbiIsICJhbXVuZC5nYWxhZW5Aa3ZlcnZhLm5vIiwgNCwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAyMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDIwfSIsIDQwLjAsICIyMDI2LTAxLTIzVDA2OjAzOjIxIl0sIFsidG9yb2xhdiIsICJUb3IgT2xhdiBBbmRlcnNlbiIsICJ0b2JhQGt2ZXJ2YS5ubyIsIDQsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMS0yM1QwNjoxODowMCJdLCBbInN5bm5lIiwgIlN5bm5lIEjDpXJzdGFkIiwgInN5bm5lQGt2ZXJ2YS5ubyIsIDQsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMS0yM1QwNjoxOToxMiJdLCBbInBpYSIsICJQaWEgSGFtbWVyIiwgInBpYS5oYW1tZXJAa3ZlcnZhLm5vIiwgNCwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAxLTIzVDA2OjI1OjAzIl0sIFsidG9yZ2VpciIsICJUb3JnZWlyIFN2YWUiLCAidG9yZ2Vpci5zdmFlQGt2ZXJ2YS5ubyIsIDQsIDIwMjYsIDAsICJ7XCJTY2FsZVwiOiA1LCBcIkluc3VsYVwiOiAxNiwgXCJJVkNcIjogMywgXCJTdHlyZW3DuHRlclwiOiA0LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMTN9IiwgNDEuMCwgIjIwMjYtMDEtMjNUMDY6NDM6MDEiXSwgWyJlcmlrIiwgIkVyaWsgRWhyZW5wb2hsIFNhbmQiLCAiZXJpay5laHJlbnBvaGwuc2FuZEBrdmVydmEubm8iLCA0LCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDI1LCBcIlNwYXJlYmFua2VuXCI6IDUsIFwiR2xvYmFsZSBha3NqZXJcIjogNSwgXCJLdmVydmEtbcO4dGVyXCI6IDV9IiwgNDAuMCwgIjIwMjYtMDEtMjNUMDg6NDQ6MDEiXSwgWyJ0b3JzdGVpbiIsICJUb3JzdGVpbiBaYWhsIiwgInR6ZUBrdmVydmEubm8iLCA0LCAyMDI2LCAwLCAie1wiTGF4VmFsb3Jpc1wiOiAyMywgXCJSYXBwb3J0ZXJpbmdcIjogMTd9IiwgNDAuMCwgIjIwMjYtMDEtMjNUMTA6Mjg6MjMiXSwgWyJwZXJuaWxsZSIsICJQZXJuaWxsZSBTa2Fyc3RlaW4iLCAicGVybmlsbGVAa3ZlcnZhLm5vIiwgNCwgMjAyNiwgMCwgIntcIkJFV2lcIjogMiwgXCJOb3JkaXNrZSBha3NqZXJcIjogMjAsIFwiR2xvYmFsZSBha3NqZXJcIjogMiwgXCJLdmVydmEtbcO4dGVyXCI6IDE2fSIsIDQwLjAsICIyMDI2LTAxLTIzVDExOjQ5OjQ0Il0sIFsiYm9yZ2UiLCAiQsO4cmdlIEtsdW5nZXJibyIsICJia0BrdmVydmEubm8iLCA0LCAyMDI2LCAwLCAie1wiSW5zdWxhXCI6IDI1LCBcIk55ZSBjYXNlIEludmVzdGVyaW5nZXJcIjogNSwgXCJLdmVydmEtbcO4dGVyXCI6IDEwfSIsIDQwLjAsICIyMDI2LTAxLTI0VDA1OjA2OjE1Il0sIFsibWF0cy5tIiwgIk1hdHMgTWFsdmlnIiwgIm1hdHMubWFsdmlnQGt2ZXJ2YS5ubyIsIDQsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMTAsIFwiSVZDXCI6IDEwLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMTIsIFwiS3ZlcnZhLW3DuHRlclwiOiA4fSIsIDQwLjAsICIyMDI2LTAxLTI1VDA3OjMwOjQxIl0sIFsiZnJvZGUiLCAiRnJvZGUgU2FuZG1hcmsiLCAiZnJvZGUuc2FuZG1hcmtAa3ZlcnZhLm5vIiwgNCwgMjAyNiwgMCwgIntcIlBlbGFnaWFcIjogMSwgXCJOdXRyaW1hclwiOiA4LCBcIkJlbmNobWFya1wiOiAxMiwgXCJMYXhWYWxvcmlzXCI6IDEzLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMiwgXCJLdmVydmEtbcO4dGVyXCI6IDR9IiwgNDAuMCwgIjIwMjYtMDEtMjZUMDI6MTA6MjUiXSwgWyJtYXRzLmciLCAiTWF0cyBHYWJyaWVsc2VuIiwgIm1hdHMuZ2FicmllbHNlbkBrdmVydmEubm8iLCA0LCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDI3LCBcIlNwYXJlYmFua2VuXCI6IDMsIFwiR2xvYmFsZSBha3NqZXJcIjogMywgXCJLdmVydmEtbcO4dGVyXCI6IDd9IiwgNDAuMCwgIjIwMjYtMDEtMjZUMDI6Mjg6NTAiXSwgWyJuaWtvbGFpIiwgIk5pa29sYWkgSmVuc2VuIiwgIm5pa29sYWkuamVuc2VuQGt2ZXJ2YS5ubyIsIDQsIDIwMjYsIDAsICJ7XCJTY2FsZVwiOiAzLCBcIkluc3VsYVwiOiAxNSwgXCJDb25jaGlsaWFcIjogMTMsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiA3LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMn0iLCA0MC4wLCAiMjAyNi0wMS0yNlQwNTowNDowOSJdLCBbInJhZ25hIiwgIlJhZ25hIEZhbGthbmdlciIsICJyYWduYS5mYWxrYW5nZXJAa3ZlcnZhLm5vIiwgNCwgMjAyNiwgMCwgIntcIlBlbGFnaWFcIjogMTAsIFwiTnV0cmltYXJcIjogMTAsIFwiSVZDXCI6IDUsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiAxMCwgXCJLdmVydmEtbcO4dGVyXCI6IDV9IiwgNDAuMCwgIjIwMjYtMDEtMjhUMDI6MDQ6MjkiXSwgWyJ0b3JvbGF2IiwgIlRvciBPbGF2IEFuZGVyc2VuIiwgInRvYmFAa3ZlcnZhLm5vIiwgNSwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAxLTMwVDA2OjA5OjEyIl0sIFsiZWlyaWsiLCAiRWlyaWsgVmFibyIsICJlaXJpay52YWJvQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjUsIFwiU2lua2FiZXJnLUhhbnNlblwiOiA1LCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMTV9IiwgNDUuMCwgIjIwMjYtMDEtMzBUMDY6MTE6NDQiXSwgWyJtb3J0ZW4iLCAiTW9ydGVuIE1qw7hlbiIsICJtYm1Aa3ZlcnZhLm5vIiwgNSwgMjAyNiwgMCwgIntcIlNjYWxlXCI6IDEyLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMTgsIFwiUmFwcG9ydGVyaW5nXCI6IDEwfSIsIDQwLjAsICIyMDI2LTAxLTMwVDA2OjEzOjE1Il0sIFsicmFnbmEiLCAiUmFnbmEgRmFsa2FuZ2VyIiwgInJhZ25hLmZhbGthbmdlckBrdmVydmEubm8iLCA1LCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAxMiwgXCJOdXRyaW1hclwiOiAxMywgXCJCZW5jaG1hcmtcIjogMTAsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiA1fSIsIDQwLjAsICIyMDI2LTAxLTMwVDA2OjEzOjQ4Il0sIFsiZXJpayIsICJFcmlrIEVocmVucG9obCBTYW5kIiwgImVyaWsuZWhyZW5wb2hsLnNhbmRAa3ZlcnZhLm5vIiwgNSwgMjAyNiwgMCwgIntcIk5vcmRpc2tlIGFrc2plclwiOiAyNiwgXCJTcGFyZWJhbmtlblwiOiA0LCBcIkdsb2JhbGUgYWtzamVyXCI6IDQsIFwiS3ZlcnZhLW3DuHRlclwiOiAyLCBcIkt2YXJ2XCI6IDR9IiwgNDAuMCwgIjIwMjYtMDEtMzBUMDY6MjU6MjgiXSwgWyJ0b3JzdGVpbiIsICJUb3JzdGVpbiBaYWhsIiwgInR6ZUBrdmVydmEubm8iLCA1LCAyMDI2LCAwLCAie1wiTGF4VmFsb3Jpc1wiOiAxNSwgXCJSYXBwb3J0ZXJpbmdcIjogMjV9IiwgNDAuMCwgIjIwMjYtMDEtMzBUMDY6Mjk6MjQiXSwgWyJwaWEiLCAiUGlhIEhhbW1lciIsICJwaWEuaGFtbWVyQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMS0zMFQwNzoxODo1OCJdLCBbImZyb2RlIiwgIkZyb2RlIFNhbmRtYXJrIiwgImZyb2RlLnNhbmRtYXJrQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDIsIFwiTnV0cmltYXJcIjogMiwgXCJCZW5jaG1hcmtcIjogMTAsIFwiTGF4VmFsb3Jpc1wiOiAyMCwgXCJOeWUgY2FzZSBTasO4bWF0XCI6IDMsIFwiS3ZlcnZhLW3DuHRlclwiOiAzfSIsIDQwLjAsICIyMDI2LTAxLTMwVDA3OjM2OjIxIl0sIFsidG9yZ2VpciIsICJUb3JnZWlyIFN2YWUiLCAidG9yZ2Vpci5zdmFlQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDcsIFwiU2NhbGVcIjogOSwgXCJJbnN1bGFcIjogMTUsIFwiSVZDXCI6IDQsIFwiU3R5cmVtw7h0ZXJcIjogNSwgXCJLdmVydmEtbcO4dGVyXCI6IDV9IiwgNDUuMCwgIjIwMjYtMDEtMzBUMDk6NDc6MDUiXSwgWyJib3JnZSIsICJCw7hyZ2UgS2x1bmdlcmJvIiwgImJrQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogMjUsIFwiQkVXaVwiOiA1LCBcIlNhbHZlc2VuICYgVGhhbXNcIjogNSwgXCJTdHlyZW3DuHRlclwiOiA1fSIsIDQwLjAsICIyMDI2LTAxLTMwVDExOjQwOjIwIl0sIFsibWF0cy5tIiwgIk1hdHMgTWFsdmlnIiwgIm1hdHMubWFsdmlnQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMTAsIFwiSVZDXCI6IDgsIFwiS2luZ2Zpc2ggQ29tcGFueVwiOiAyLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMjB9IiwgNDAuMCwgIjIwMjYtMDEtMzBUMTM6MjI6MzEiXSwgWyJvbGF2IiwgIk9sYXYgSG9sc3QgRHlybmVzIiwgIm9sYXYuaG9sc3QuZHlybmVzQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogMjIsIFwiQ29uY2hpbGlhXCI6IDksIFwiTnllIGNhc2UgU2rDuG1hdFwiOiAyLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMX0iLCAzNC4wLCAiMjAyNi0wMS0zMVQwNDoxNTo0OCJdLCBbInNpbWVuIiwgIlNpbWVuIE5pbHNlbiIsICJzb25Aa3ZlcnZhLm5vIiwgNSwgMjAyNiwgMCwgIntcIkJFV2lcIjogOCwgXCJTYWx2ZXNlbiAmIFRoYW1zXCI6IDgsIFwiTnllIGNhc2UgSW52ZXN0ZXJpbmdlclwiOiA4LCBcIlJhcHBvcnRlcmluZ1wiOiA4LCBcIkVuZHVyXCI6IDh9IiwgNDAuMCwgIjIwMjYtMDEtMzFUMTE6NDM6MzciXSwgWyJzeW5uZSIsICJTeW5uZSBIw6Vyc3RhZCIsICJzeW5uZUBrdmVydmEubm8iLCA1LCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogMzV9IiwgMzUuMCwgIjIwMjYtMDItMDJUMDE6NTI6MTYiXSwgWyJhbXVuZCIsICJBbXVuZCBHYWxhZW4iLCAiYW11bmQuZ2FsYWVuQGt2ZXJ2YS5ubyIsIDUsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMi0wMlQwOTo1Nzo0MCJdLCBbInRvcm9sYXYiLCAiVG9yIE9sYXYgQW5kZXJzZW4iLCAidG9iYUBrdmVydmEubm8iLCA2LCAyMDI2LCAwLCAie1wiQ29uY2hpbGlhXCI6IDgsIFwiUmVnbnNrYXBcIjogMzJ9IiwgNDAuMCwgIjIwMjYtMDItMDZUMDY6MDI6MTIiXSwgWyJlaXJpayIsICJFaXJpayBWYWJvIiwgImVpcmlrLnZhYm9Aa3ZlcnZhLm5vIiwgNiwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAzMywgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDcsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiAxMCwgXCJLdmVydmEtbcO4dGVyXCI6IDV9IiwgNTUuMCwgIjIwMjYtMDItMDZUMDY6MDU6NDYiXSwgWyJlcmlrIiwgIkVyaWsgRWhyZW5wb2hsIFNhbmQiLCAiZXJpay5laHJlbnBvaGwuc2FuZEBrdmVydmEubm8iLCA2LCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDI0LCBcIlNwYXJlYmFua2VuXCI6IDQsIFwiR2xvYmFsZSBha3NqZXJcIjogNCwgXCJSYXBwb3J0ZXJpbmdcIjogMywgXCJLdmVydmEtbcO4dGVyXCI6IDEsIFwiS3ZhcnZcIjogNH0iLCA0MC4wLCAiMjAyNi0wMi0wNlQwNjoxMDowNSJdLCBbInBpYSIsICJQaWEgSGFtbWVyIiwgInBpYS5oYW1tZXJAa3ZlcnZhLm5vIiwgNiwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAyLTA2VDA2OjI0OjU1Il0sIFsic2ltZW4iLCAiU2ltZW4gTmlsc2VuIiwgInNvbkBrdmVydmEubm8iLCA2LCAyMDI2LCAwLCAie1wiQkVXaVwiOiA4LCBcIlNhbHZlc2VuICYgVGhhbXNcIjogMTYsIFwiTnllIGNhc2UgSW52ZXN0ZXJpbmdlclwiOiA1LCBcIlJhcHBvcnRlcmluZ1wiOiAzLCBcIkVuZHVyXCI6IDh9IiwgNDAuMCwgIjIwMjYtMDItMDZUMDc6NDY6NDciXSwgWyJ0b3JzdGVpbiIsICJUb3JzdGVpbiBaYWhsIiwgInR6ZUBrdmVydmEubm8iLCA2LCAyMDI2LCAwLCAie1wiTGF4VmFsb3Jpc1wiOiAzMCwgXCJSYXBwb3J0ZXJpbmdcIjogMTB9IiwgNDAuMCwgIjIwMjYtMDItMDZUMDg6MzI6MjQiXSwgWyJhbXVuZCIsICJBbXVuZCBHYWxhZW4iLCAiYW11bmQuZ2FsYWVuQGt2ZXJ2YS5ubyIsIDYsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMi0wNlQwODozNjo0OSJdLCBbIm1vcnRlbiIsICJNb3J0ZW4gTWrDuGVuIiwgIm1ibUBrdmVydmEubm8iLCA2LCAyMDI2LCAwLCAie1wiU2NhbGVcIjogMjAsIFwiUmFwcG9ydGVyaW5nXCI6IDIwfSIsIDQwLjAsICIyMDI2LTAyLTA2VDA4OjQ1OjIxIl0sIFsidG9yZ2VpciIsICJUb3JnZWlyIFN2YWUiLCAidG9yZ2Vpci5zdmFlQGt2ZXJ2YS5ubyIsIDYsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDIsIFwiU2NhbGVcIjogMTEsIFwiSW5zdWxhXCI6IDksIFwiSVZDXCI6IDYsIFwiU3R5cmVtw7h0ZXJcIjogOSwgXCJLdmVydmEtbcO4dGVyXCI6IDUsIFwiS3ZhcnZcIjogM30iLCA0NS4wLCAiMjAyNi0wMi0wNlQxMDo0MTo0NiJdLCBbImJvcmdlIiwgIkLDuHJnZSBLbHVuZ2VyYm8iLCAiYmtAa3ZlcnZhLm5vIiwgNiwgMjAyNiwgMCwgIntcIkluc3VsYVwiOiAyNSwgXCJTYWx2ZXNlbiAmIFRoYW1zXCI6IDYsIFwiU3R5cmVtw7h0ZXJcIjogNCwgXCJFbmR1clwiOiA1fSIsIDQwLjAsICIyMDI2LTAyLTA2VDEzOjE5OjQ0Il0sIFsibWFnbnVzIiwgIk1hZ251cyBEeWJ2YWQiLCAibWRAa3ZlcnZhLm5vIiwgNiwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAyMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDIsIFwiTnllIGNhc2UgTGFrc2VvcHBkcmV0dFwiOiAxMCwgXCJMYXhWYWxvcmlzXCI6IDMsIFwiUmFwcG9ydGVyaW5nXCI6IDIsIFwiS3ZlcnZhLW3DuHRlclwiOiAzfSIsIDQwLjAsICIyMDI2LTAyLTA4VDAyOjQ5OjM3Il0sIFsiZnJvZGUiLCAiRnJvZGUgU2FuZG1hcmsiLCAiZnJvZGUuc2FuZG1hcmtAa3ZlcnZhLm5vIiwgNiwgMjAyNiwgMCwgIntcIkJlbmNobWFya1wiOiAxMiwgXCJJVkNcIjogMiwgXCJMYXhWYWxvcmlzXCI6IDE2LCBcIk55ZSBjYXNlIFNqw7htYXRcIjogNCwgXCJLdmVydmEtbcO4dGVyXCI6IDZ9IiwgNDAuMCwgIjIwMjYtMDItMDlUMDI6MTE6MzAiXSwgWyJyYWduYSIsICJSYWduYSBGYWxrYW5nZXIiLCAicmFnbmEuZmFsa2FuZ2VyQGt2ZXJ2YS5ubyIsIDYsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDEwLCBcIk51dHJpbWFyXCI6IDEwLCBcIkJlbmNobWFya1wiOiA1LCBcIklWQ1wiOiAxMCwgXCJSYXBwb3J0ZXJpbmdcIjogNX0iLCA0MC4wLCAiMjAyNi0wMi0wOVQwMjo1Njo0MCJdLCBbInN5bm5lIiwgIlN5bm5lIEjDpXJzdGFkIiwgInN5bm5lQGt2ZXJ2YS5ubyIsIDYsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiAzNH0iLCAzNC4wLCAiMjAyNi0wMi0wOVQwMjo1ODowOCJdLCBbIm9sYXYiLCAiT2xhdiBIb2xzdCBEeXJuZXMiLCAib2xhdi5ob2xzdC5keXJuZXNAa3ZlcnZhLm5vIiwgNiwgMjAyNiwgMCwgIntcIkluc3VsYVwiOiAxNSwgXCJDb25jaGlsaWFcIjogNSwgXCJLdmVydmEtbcO4dGVyXCI6IDJ9IiwgMjIuMCwgIjIwMjYtMDItMTJUMDk6MTA6MzMiXSwgWyJwZXJuaWxsZSIsICJQZXJuaWxsZSBTa2Fyc3RlaW4iLCAicGVybmlsbGVAa3ZlcnZhLm5vIiwgNiwgMjAyNiwgMCwgIntcIkJFV2lcIjogNiwgXCJOb3JkaXNrZSBha3NqZXJcIjogMjIsIFwiR2xvYmFsZSBha3NqZXJcIjogMiwgXCJTdHlyZW3DuHRlclwiOiA2LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogNH0iLCA0MC4wLCAiMjAyNi0wMi0xNVQwNzoxOTowNiJdLCBbInRvcm9sYXYiLCAiVG9yIE9sYXYgQW5kZXJzZW4iLCAidG9iYUBrdmVydmEubm8iLCA3LCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogNDB9IiwgNDAuMCwgIjIwMjYtMDItMTNUMDY6MTA6MjciXSwgWyJvbGF2IiwgIk9sYXYgSG9sc3QgRHlybmVzIiwgIm9sYXYuaG9sc3QuZHlybmVzQGt2ZXJ2YS5ubyIsIDcsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogMjcsIFwiQ29uY2hpbGlhXCI6IDE1fSIsIDQyLjAsICIyMDI2LTAyLTEzVDA2OjMyOjMzIl0sIFsicGlhIiwgIlBpYSBIYW1tZXIiLCAicGlhLmhhbW1lckBrdmVydmEubm8iLCA3LCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogNDB9IiwgNDAuMCwgIjIwMjYtMDItMTNUMDY6Mzk6NDAiXSwgWyJtYWdudXMiLCAiTWFnbnVzIER5YnZhZCIsICJtZEBrdmVydmEubm8iLCA3LCAyMDI2LCAwLCAie1wiU2FsTWFyXCI6IDEwLCBcIk51dHJpbWFyXCI6IDEwLCBcIlJhcHBvcnRlcmluZ1wiOiA1LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogNSwgXCJFbmR1clwiOiAxMH0iLCA0MC4wLCAiMjAyNi0wMi0xM1QwNjo0NToyMiJdLCBbInRvcnN0ZWluIiwgIlRvcnN0ZWluIFphaGwiLCAidHplQGt2ZXJ2YS5ubyIsIDcsIDIwMjYsIDAsICJ7XCJMYXhWYWxvcmlzXCI6IDI5LCBcIlJhcHBvcnRlcmluZ1wiOiAxMX0iLCA0MC4wLCAiMjAyNi0wMi0xM1QwNjo0ODowNyJdLCBbImFtdW5kIiwgIkFtdW5kIEdhbGFlbiIsICJhbXVuZC5nYWxhZW5Aa3ZlcnZhLm5vIiwgNywgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAyMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDIwfSIsIDQwLjAsICIyMDI2LTAyLTEzVDA2OjU2OjEzIl0sIFsibmlrb2xhaSIsICJOaWtvbGFpIEplbnNlbiIsICJuaWtvbGFpLmplbnNlbkBrdmVydmEubm8iLCA3LCAyMDI2LCAwLCAie1wiSW5zdWxhXCI6IDM1LCBcIkNvbmNoaWxpYVwiOiA1fSIsIDQwLjAsICIyMDI2LTAyLTEzVDA3OjMwOjE2Il0sIFsicGVybmlsbGUiLCAiUGVybmlsbGUgU2thcnN0ZWluIiwgInBlcm5pbGxlQGt2ZXJ2YS5ubyIsIDcsIDIwMjYsIDAsICJ7XCJCRVdpXCI6IDEyLCBcIk5vcmRpc2tlIGFrc2plclwiOiA0LCBcIlN0eXJlbcO4dGVyXCI6IDE2LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogNiwgXCJLdmFydlwiOiAyfSIsIDQwLjAsICIyMDI2LTAyLTEzVDA3OjQyOjU2Il0sIFsibWF0cy5nIiwgIk1hdHMgR2FicmllbHNlbiIsICJtYXRzLmdhYnJpZWxzZW5Aa3ZlcnZhLm5vIiwgNywgMjAyNiwgMCwgIntcIk5vcmRpc2tlIGFrc2plclwiOiAzMCwgXCJTcGFyZWJhbmtlblwiOiA0LCBcIkdsb2JhbGUgYWtzamVyXCI6IDQsIFwiS3ZlcnZhLW3DuHRlclwiOiAyfSIsIDQwLjAsICIyMDI2LTAyLTEzVDA3OjUyOjA5Il0sIFsiZXJpayIsICJFcmlrIEVocmVucG9obCBTYW5kIiwgImVyaWsuZWhyZW5wb2hsLnNhbmRAa3ZlcnZhLm5vIiwgNywgMjAyNiwgMCwgIntcIk5vcmRpc2tlIGFrc2plclwiOiAzMCwgXCJTcGFyZWJhbmtlblwiOiA0LCBcIkdsb2JhbGUgYWtzamVyXCI6IDQsIFwiS3ZlcnZhLW3DuHRlclwiOiAyfSIsIDQwLjAsICIyMDI2LTAyLTEzVDA3OjU2OjI3Il0sIFsicmFnbmEiLCAiUmFnbmEgRmFsa2FuZ2VyIiwgInJhZ25hLmZhbGthbmdlckBrdmVydmEubm8iLCA3LCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAxMCwgXCJOdXRyaW1hclwiOiAxNSwgXCJCZW5jaG1hcmtcIjogMTAsIFwiSVZDXCI6IDV9IiwgNDAuMCwgIjIwMjYtMDItMTNUMTA6MDI6NTMiXSwgWyJlaXJpayIsICJFaXJpayBWYWJvIiwgImVpcmlrLnZhYm9Aa3ZlcnZhLm5vIiwgNywgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAyMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDIwLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMTB9IiwgNTAuMCwgIjIwMjYtMDItMTNUMTE6MjY6NDQiXSwgWyJzeW5uZSIsICJTeW5uZSBIw6Vyc3RhZCIsICJzeW5uZUBrdmVydmEubm8iLCA3LCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogNDB9IiwgNDAuMCwgIjIwMjYtMDItMTZUMDI6MDU6MjMiXSwgWyJzaW1lbiIsICJTaW1lbiBOaWxzZW4iLCAic29uQGt2ZXJ2YS5ubyIsIDcsIDIwMjYsIDAsICJ7XCJCRVdpXCI6IDYsIFwiU2FsdmVzZW4gJiBUaGFtc1wiOiAxMCwgXCJSYXBwb3J0ZXJpbmdcIjogNCwgXCJFbmR1clwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMi0xNlQwMzowNDowNCJdLCBbInRvcnN0ZWluIiwgIlRvcnN0ZWluIFphaGwiLCAidHplQGt2ZXJ2YS5ubyIsIDgsIDIwMjYsIDAsICJ7XCJMYXhWYWxvcmlzXCI6IDI1LCBcIlJhcHBvcnRlcmluZ1wiOiAxNX0iLCA0MC4wLCAiMjAyNi0wMi0yMFQwNjowODowOCJdLCBbIm1vcnRlbiIsICJNb3J0ZW4gTWrDuGVuIiwgIm1ibUBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiU2NhbGVcIjogMjUsIFwiUmFwcG9ydGVyaW5nXCI6IDE1fSIsIDQwLjAsICIyMDI2LTAyLTIwVDA2OjEwOjEwIl0sIFsiYW11bmQiLCAiQW11bmQgR2FsYWVuIiwgImFtdW5kLmdhbGFlbkBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiU2FsTWFyXCI6IDIwLCBcIlNpbmthYmVyZy1IYW5zZW5cIjogMjB9IiwgNDAuMCwgIjIwMjYtMDItMjBUMDY6MzM6MzIiXSwgWyJlaXJpayIsICJFaXJpayBWYWJvIiwgImVpcmlrLnZhYm9Aa3ZlcnZhLm5vIiwgOCwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAyNSwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDI1LCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMTB9IiwgNjAuMCwgIjIwMjYtMDItMjBUMDY6Mzg6NTgiXSwgWyJvbGF2IiwgIk9sYXYgSG9sc3QgRHlybmVzIiwgIm9sYXYuaG9sc3QuZHlybmVzQGt2ZXJ2YS5ubyIsIDgsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogNDAsIFwiQ29uY2hpbGlhXCI6IDEwfSIsIDUwLjAsICIyMDI2LTAyLTIwVDA2OjQ0OjE1Il0sIFsicGlhIiwgIlBpYSBIYW1tZXIiLCAicGlhLmhhbW1lckBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogNDB9IiwgNDAuMCwgIjIwMjYtMDItMjBUMDY6NDY6NDkiXSwgWyJlcmlrIiwgIkVyaWsgRWhyZW5wb2hsIFNhbmQiLCAiZXJpay5laHJlbnBvaGwuc2FuZEBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDMwLCBcIlNwYXJlYmFua2VuXCI6IDQsIFwiR2xvYmFsZSBha3NqZXJcIjogNCwgXCJLdmVydmEtbcO4dGVyXCI6IDJ9IiwgNDAuMCwgIjIwMjYtMDItMjBUMDc6MTk6NTAiXSwgWyJtYXRzLmciLCAiTWF0cyBHYWJyaWVsc2VuIiwgIm1hdHMuZ2FicmllbHNlbkBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDMxLCBcIlNwYXJlYmFua2VuXCI6IDUsIFwiR2xvYmFsZSBha3NqZXJcIjogMywgXCJLdmVydmEtbcO4dGVyXCI6IDF9IiwgNDAuMCwgIjIwMjYtMDItMjBUMDc6MjM6NTUiXSwgWyJ0b3JvbGF2IiwgIlRvciBPbGF2IEFuZGVyc2VuIiwgInRvYmFAa3ZlcnZhLm5vIiwgOCwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDM1fSIsIDM1LjAsICIyMDI2LTAyLTIwVDA3OjUxOjAyIl0sIFsic2ltZW4iLCAiU2ltZW4gTmlsc2VuIiwgInNvbkBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiQkVXaVwiOiAxMCwgXCJTYWx2ZXNlbiAmIFRoYW1zXCI6IDEyLCBcIk55ZSBjYXNlIEludmVzdGVyaW5nZXJcIjogNSwgXCJSYXBwb3J0ZXJpbmdcIjogOCwgXCJFbmR1clwiOiA1fSIsIDQwLjAsICIyMDI2LTAyLTIyVDE1OjE0OjIwIl0sIFsidG9yZ2VpciIsICJUb3JnZWlyIFN2YWUiLCAidG9yZ2Vpci5zdmFlQGt2ZXJ2YS5ubyIsIDgsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDQsIFwiU2NhbGVcIjogMTgsIFwiSW5zdWxhXCI6IDgsIFwiSVZDXCI6IDMsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiA1LCBcIlN0eXJlbcO4dGVyXCI6IDUsIFwiS3ZlcnZhLW3DuHRlclwiOiA1fSIsIDQ4LjAsICIyMDI2LTAyLTIzVDAxOjQyOjE2Il0sIFsicmFnbmEiLCAiUmFnbmEgRmFsa2FuZ2VyIiwgInJhZ25hLmZhbGthbmdlckBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAxMCwgXCJOdXRyaW1hclwiOiAxNSwgXCJCZW5jaG1hcmtcIjogMTAsIFwiSVZDXCI6IDV9IiwgNDAuMCwgIjIwMjYtMDItMjNUMDI6NTQ6MjkiXSwgWyJzeW5uZSIsICJTeW5uZSBIw6Vyc3RhZCIsICJzeW5uZUBrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogNDB9IiwgNDAuMCwgIjIwMjYtMDItMjNUMDQ6Mjc6MDMiXSwgWyJmcm9kZSIsICJGcm9kZSBTYW5kbWFyayIsICJmcm9kZS5zYW5kbWFya0BrdmVydmEubm8iLCA4LCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAzLCBcIk51dHJpbWFyXCI6IDMsIFwiQmVuY2htYXJrXCI6IDEwLCBcIklWQ1wiOiAxLCBcIkxheFZhbG9yaXNcIjogMTAsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiAyLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogNH0iLCAzMy4wLCAiMjAyNi0wMi0yM1QxMDoyOTo1NiJdLCBbInBpYSIsICJQaWEgSGFtbWVyIiwgInBpYS5oYW1tZXJAa3ZlcnZhLm5vIiwgOSwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAyLTI3VDA2OjExOjMyIl0sIFsiYW11bmQiLCAiQW11bmQgR2FsYWVuIiwgImFtdW5kLmdhbGFlbkBrdmVydmEubm8iLCA5LCAyMDI2LCAwLCAie1wiU2FsTWFyXCI6IDIwLCBcIlNpbmthYmVyZy1IYW5zZW5cIjogMjB9IiwgNDAuMCwgIjIwMjYtMDItMjdUMDY6MTM6NTEiXSwgWyJ0b3JvbGF2IiwgIlRvciBPbGF2IEFuZGVyc2VuIiwgInRvYmFAa3ZlcnZhLm5vIiwgOSwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAyLTI3VDA2OjE3OjAyIl0sIFsidG9yc3RlaW4iLCAiVG9yc3RlaW4gWmFobCIsICJ0emVAa3ZlcnZhLm5vIiwgOSwgMjAyNiwgMCwgIntcIkxheFZhbG9yaXNcIjogMjksIFwiUmFwcG9ydGVyaW5nXCI6IDExfSIsIDQwLjAsICIyMDI2LTAyLTI3VDA2OjE4OjU0Il0sIFsicmFnbmEiLCAiUmFnbmEgRmFsa2FuZ2VyIiwgInJhZ25hLmZhbGthbmdlckBrdmVydmEubm8iLCA5LCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAxMywgXCJOdXRyaW1hclwiOiAxMiwgXCJCZW5jaG1hcmtcIjogNSwgXCJSYXBwb3J0ZXJpbmdcIjogMTB9IiwgNDAuMCwgIjIwMjYtMDItMjdUMDY6Mjc6MTYiXSwgWyJzeW5uZSIsICJTeW5uZSBIw6Vyc3RhZCIsICJzeW5uZUBrdmVydmEubm8iLCA5LCAyMDI2LCAwLCAie1wiUmVnbnNrYXBcIjogNDB9IiwgNDAuMCwgIjIwMjYtMDItMjdUMDY6Mjk6MTYiXSwgWyJwZXJuaWxsZSIsICJQZXJuaWxsZSBTa2Fyc3RlaW4iLCAicGVybmlsbGVAa3ZlcnZhLm5vIiwgOSwgMjAyNiwgMCwgIntcIk5vcmRpc2tlIGFrc2plclwiOiAzMSwgXCJHbG9iYWxlIGFrc2plclwiOiAyLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogN30iLCA0MC4wLCAiMjAyNi0wMi0yN1QwNjo0MDowMCJdLCBbIm1hdHMubSIsICJNYXRzIE1hbHZpZyIsICJtYXRzLm1hbHZpZ0BrdmVydmEubm8iLCA5LCAyMDI2LCAwLCAie1wiU2FsTWFyXCI6IDgsIFwiSVZDXCI6IDgsIFwiTnllIGNhc2UgU2rDuG1hdFwiOiAyNH0iLCA0MC4wLCAiMjAyNi0wMi0yN1QwNzowMjo0MyJdLCBbIm1vcnRlbiIsICJNb3J0ZW4gTWrDuGVuIiwgIm1ibUBrdmVydmEubm8iLCA5LCAyMDI2LCAwLCAie1wiU2NhbGVcIjogMzQsIFwiUmFwcG9ydGVyaW5nXCI6IDZ9IiwgNDAuMCwgIjIwMjYtMDItMjdUMDg6MDE6MTQiXSwgWyJlaXJpayIsICJFaXJpayBWYWJvIiwgImVpcmlrLnZhYm9Aa3ZlcnZhLm5vIiwgOSwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAxMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDM1LCBcIk55ZSBjYXNlIFNqw7htYXRcIjogNX0iLCA1MC4wLCAiMjAyNi0wMi0yN1QwODo1MToyNiJdLCBbInRvcmdlaXIiLCAiVG9yZ2VpciBTdmFlIiwgInRvcmdlaXIuc3ZhZUBrdmVydmEubm8iLCA5LCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiA1LCBcIlNjYWxlXCI6IDEwLCBcIkluc3VsYVwiOiA5LCBcIklWQ1wiOiA1LCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMywgXCJLdmVydmEtbcO4dGVyXCI6IDQsIFwiS3ZhcnZcIjogNH0iLCA0MC4wLCAiMjAyNi0wMi0yN1QwOTo0OTowOSJdLCBbImVyaWsiLCAiRXJpayBFaHJlbnBvaGwgU2FuZCIsICJlcmlrLmVocmVucG9obC5zYW5kQGt2ZXJ2YS5ubyIsIDksIDIwMjYsIDAsICJ7XCJOb3JkaXNrZSBha3NqZXJcIjogMzAsIFwiU3BhcmViYW5rZW5cIjogNCwgXCJHbG9iYWxlIGFrc2plclwiOiA0LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMn0iLCA0MC4wLCAiMjAyNi0wMy0wMlQwMjowMDo1NyJdLCBbIm9sYXYiLCAiT2xhdiBIb2xzdCBEeXJuZXMiLCAib2xhdi5ob2xzdC5keXJuZXNAa3ZlcnZhLm5vIiwgOSwgMjAyNiwgMCwgIntcIkluc3VsYVwiOiAxOCwgXCJDb25jaGlsaWFcIjogM30iLCAyMS4wLCAiMjAyNi0wMy0wMlQwNDo0ODo0NSJdLCBbInNpbWVuIiwgIlNpbWVuIE5pbHNlbiIsICJzb25Aa3ZlcnZhLm5vIiwgOSwgMjAyNiwgMCwgIntcIkJFV2lcIjogOCwgXCJOeWUgY2FzZSBJbnZlc3RlcmluZ2VyXCI6IDIyLCBcIlJhcHBvcnRlcmluZ1wiOiA2LCBcIkVuZHVyXCI6IDR9IiwgNDAuMCwgIjIwMjYtMDMtMDJUMDc6MjI6MDMiXSwgWyJtYWdudXMiLCAiTWFnbnVzIER5YnZhZCIsICJtZEBrdmVydmEubm8iLCA5LCAyMDI2LCAwLCAie1wiU2FsTWFyXCI6IDE1LCBcIlNpbmthYmVyZy1IYW5zZW5cIjogNSwgXCJOdXRyaW1hclwiOiAxMCwgXCJSZWduc2thcFwiOiAyLCBcIlJhcHBvcnRlcmluZ1wiOiA1LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogM30iLCA0MC4wLCAiMjAyNi0wMy0wMlQxMzo0OTo0NyJdLCBbIm1vcnRlbiIsICJNb3J0ZW4gTWrDuGVuIiwgIm1ibUBrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIlNjYWxlXCI6IDEzLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMjd9IiwgNDAuMCwgIjIwMjYtMDMtMDZUMDY6MDI6MDEiXSwgWyJuaWtvbGFpIiwgIk5pa29sYWkgSmVuc2VuIiwgIm5pa29sYWkuamVuc2VuQGt2ZXJ2YS5ubyIsIDEwLCAyMDI2LCAwLCAie1wiSW5zdWxhXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAzLTA2VDA2OjA0OjM2Il0sIFsidG9yb2xhdiIsICJUb3IgT2xhdiBBbmRlcnNlbiIsICJ0b2JhQGt2ZXJ2YS5ubyIsIDEwLCAyMDI2LCAwLCAie1wiUmFwcG9ydGVyaW5nXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAzLTA2VDA2OjEwOjMyIl0sIFsicGlhIiwgIlBpYSBIYW1tZXIiLCAicGlhLmhhbW1lckBrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAzLTA2VDA2OjExOjQ4Il0sIFsic3lubmUiLCAiU3lubmUgSMOlcnN0YWQiLCAic3lubmVAa3ZlcnZhLm5vIiwgMTAsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMy0wNlQwNjoxMjowMiJdLCBbImVpcmlrIiwgIkVpcmlrIFZhYm8iLCAiZWlyaWsudmFib0BrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAyMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDIwLCBcIk55ZSBjYXNlIExha3Nlb3BwZHJldHRcIjogMTV9IiwgNTUuMCwgIjIwMjYtMDMtMDZUMDY6MjY6MjkiXSwgWyJtYWdudXMiLCAiTWFnbnVzIER5YnZhZCIsICJtZEBrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAxMCwgXCJTaW5rYWJlcmctSGFuc2VuXCI6IDEwLCBcIk55ZSBjYXNlIExha3Nlb3BwZHJldHRcIjogMTAsIFwiU2NhbGVcIjogMywgXCJSZWduc2thcFwiOiAzLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogNH0iLCA0MC4wLCAiMjAyNi0wMy0wNlQwNjozMTo1MyJdLCBbInBlcm5pbGxlIiwgIlBlcm5pbGxlIFNrYXJzdGVpbiIsICJwZXJuaWxsZUBrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIkJFV2lcIjogMiwgXCJOb3JkaXNrZSBha3NqZXJcIjogMzAsIFwiR2xvYmFsZSBha3NqZXJcIjogMiwgXCJLdmVydmEtbcO4dGVyXCI6IDZ9IiwgNDAuMCwgIjIwMjYtMDMtMDZUMDY6NDE6MzkiXSwgWyJ0b3JzdGVpbiIsICJUb3JzdGVpbiBaYWhsIiwgInR6ZUBrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIkxheFZhbG9yaXNcIjogMjAsIFwiUmFwcG9ydGVyaW5nXCI6IDIwfSIsIDQwLjAsICIyMDI2LTAzLTA2VDA3OjE0OjI4Il0sIFsiYm9yZ2UiLCAiQsO4cmdlIEtsdW5nZXJibyIsICJia0BrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIkluc3VsYVwiOiAxMCwgXCJOeWUgY2FzZSBJbnZlc3RlcmluZ2VyXCI6IDMwfSIsIDQwLjAsICIyMDI2LTAzLTA2VDA3OjMwOjIwIl0sIFsib2xhdiIsICJPbGF2IEhvbHN0IER5cm5lcyIsICJvbGF2LmhvbHN0LmR5cm5lc0BrdmVydmEubm8iLCAxMCwgMjAyNiwgMCwgIntcIkluc3VsYVwiOiAxMywgXCJDb25jaGlsaWFcIjogOSwgXCJOeWUgY2FzZSBTasO4bWF0XCI6IDE4fSIsIDQwLjAsICIyMDI2LTAzLTA2VDA3OjM0OjM5Il0sIFsiZnJvZGUiLCAiRnJvZGUgU2FuZG1hcmsiLCAiZnJvZGUuc2FuZG1hcmtAa3ZlcnZhLm5vIiwgMTAsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDEwLCBcIkJlbmNobWFya1wiOiAxMCwgXCJJVkNcIjogMiwgXCJMYXhWYWxvcmlzXCI6IDEwLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogNSwgXCJLdmVydmEtbcO4dGVyXCI6IDN9IiwgNDAuMCwgIjIwMjYtMDMtMDZUMDg6NDI6MjgiXSwgWyJyYWduYSIsICJSYWduYSBGYWxrYW5nZXIiLCAicmFnbmEuZmFsa2FuZ2VyQGt2ZXJ2YS5ubyIsIDEwLCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiA1LCBcIk51dHJpbWFyXCI6IDEwLCBcIk55ZSBjYXNlIFNqw7htYXRcIjogMjV9IiwgNDAuMCwgIjIwMjYtMDMtMDZUMDk6MjE6MDAiXSwgWyJzaW1lbiIsICJTaW1lbiBOaWxzZW4iLCAic29uQGt2ZXJ2YS5ubyIsIDEwLCAyMDI2LCAwLCAie1wiQkVXaVwiOiA0LCBcIlNhbHZlc2VuICYgVGhhbXNcIjogNCwgXCJOeWUgY2FzZSBJbnZlc3RlcmluZ2VyXCI6IDI2LCBcIlJhcHBvcnRlcmluZ1wiOiAzLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogM30iLCA0MC4wLCAiMjAyNi0wMy0wOVQwMzo1NTowMyJdLCBbImFtdW5kIiwgIkFtdW5kIEdhbGFlbiIsICJhbXVuZC5nYWxhZW5Aa3ZlcnZhLm5vIiwgMTAsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMy0wOVQwOTo1MzoyMSJdLCBbImVyaWsiLCAiRXJpayBFaHJlbnBvaGwgU2FuZCIsICJlcmlrLmVocmVucG9obC5zYW5kQGt2ZXJ2YS5ubyIsIDEwLCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDMwLCBcIlNwYXJlYmFua2VuXCI6IDUsIFwiR2xvYmFsZSBha3NqZXJcIjogNX0iLCA0MC4wLCAiMjAyNi0wMy0xMlQwMzozNTozOCJdLCBbInRvcnN0ZWluIiwgIlRvcnN0ZWluIFphaGwiLCAidHplQGt2ZXJ2YS5ubyIsIDExLCAyMDI2LCAwLCAie1wiTGF4VmFsb3Jpc1wiOiAyMCwgXCJSYXBwb3J0ZXJpbmdcIjogMjB9IiwgNDAuMCwgIjIwMjYtMDMtMTNUMDc6MDE6MzYiXSwgWyJ0b3JvbGF2IiwgIlRvciBPbGF2IEFuZGVyc2VuIiwgInRvYmFAa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMy0xM1QwNzowOTowNCJdLCBbImFtdW5kIiwgIkFtdW5kIEdhbGFlbiIsICJhbXVuZC5nYWxhZW5Aa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMy0xM1QwNzo0NzowNiJdLCBbInBpYSIsICJQaWEgSGFtbWVyIiwgInBpYS5oYW1tZXJAa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMy0xM1QwODoxNjo1NiJdLCBbInJhZ25hIiwgIlJhZ25hIEZhbGthbmdlciIsICJyYWduYS5mYWxrYW5nZXJAa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDE1LCBcIk51dHJpbWFyXCI6IDE1LCBcIkJlbmNobWFya1wiOiAyLCBcIklWQ1wiOiA4fSIsIDQwLjAsICIyMDI2LTAzLTEzVDA4OjQxOjU5Il0sIFsiZXJpayIsICJFcmlrIEVocmVucG9obCBTYW5kIiwgImVyaWsuZWhyZW5wb2hsLnNhbmRAa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJOb3JkaXNrZSBha3NqZXJcIjogMjgsIFwiU3BhcmViYW5rZW5cIjogNSwgXCJHbG9iYWxlIGFrc2plclwiOiA1LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMn0iLCA0MC4wLCAiMjAyNi0wMy0xM1QwOToyMjo1OCJdLCBbInBlcm5pbGxlIiwgIlBlcm5pbGxlIFNrYXJzdGVpbiIsICJwZXJuaWxsZUBrdmVydmEubm8iLCAxMSwgMjAyNiwgMCwgIntcIkJFV2lcIjogMiwgXCJOb3JkaXNrZSBha3NqZXJcIjogMjYsIFwiR2xvYmFsZSBha3NqZXJcIjogMiwgXCJTdHlyZW3DuHRlclwiOiA0LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogNCwgXCJLdmFydlwiOiAyfSIsIDQwLjAsICIyMDI2LTAzLTEzVDA5OjM3OjI1Il0sIFsidG9yZ2VpciIsICJUb3JnZWlyIFN2YWUiLCAidG9yZ2Vpci5zdmFlQGt2ZXJ2YS5ubyIsIDExLCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAyLCBcIlNjYWxlXCI6IDUsIFwiSW5zdWxhXCI6IDcsIFwiQ29uY2hpbGlhXCI6IDIsIFwiSVZDXCI6IDIsIFwiTnllIGNhc2UgSW52ZXN0ZXJpbmdlclwiOiAyLCBcIlN0eXJlbcO4dGVyXCI6IDE0LCBcIlJlZ25za2FwXCI6IDQsIFwiUmFwcG9ydGVyaW5nXCI6IDQsIFwiS3ZlcnZhLW3DuHRlclwiOiA0fSIsIDQ2LjAsICIyMDI2LTAzLTEzVDExOjA2OjU3Il0sIFsic3lubmUiLCAiU3lubmUgSMOlcnN0YWQiLCAic3lubmVAa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMy0xNlQwMzowMToyNyJdLCBbInNpbWVuIiwgIlNpbWVuIE5pbHNlbiIsICJzb25Aa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJCRVdpXCI6IDUsIFwiU2FsdmVzZW4gJiBUaGFtc1wiOiA0LCBcIk55ZSBjYXNlIEludmVzdGVyaW5nZXJcIjogMjYsIFwiRW5kdXJcIjogNX0iLCA0MC4wLCAiMjAyNi0wMy0xNlQwODoyMTo1NSJdLCBbIm9sYXYiLCAiT2xhdiBIb2xzdCBEeXJuZXMiLCAib2xhdi5ob2xzdC5keXJuZXNAa3ZlcnZhLm5vIiwgMTEsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogMjMsIFwiQ29uY2hpbGlhXCI6IDEwLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogMn0iLCAzNS4wLCAiMjAyNi0wMy0xNlQxNjowNzo1MSJdLCBbIm1hdHMuZyIsICJNYXRzIEdhYnJpZWxzZW4iLCAibWF0cy5nYWJyaWVsc2VuQGt2ZXJ2YS5ubyIsIDEyLCAyMDI2LCAwLCAie1wiTm9yZGlza2UgYWtzamVyXCI6IDI5LCBcIlNwYXJlYmFua2VuXCI6IDMsIFwiR2xvYmFsZSBha3NqZXJcIjogMywgXCJSYXBwb3J0ZXJpbmdcIjogNCwgXCJLdmVydmEtbcO4dGVyXCI6IDF9IiwgNDAuMCwgIjIwMjYtMDMtMjBUMDc6MDg6MjEiXSwgWyJ0b3JzdGVpbiIsICJUb3JzdGVpbiBaYWhsIiwgInR6ZUBrdmVydmEubm8iLCAxMiwgMjAyNiwgMCwgIntcIkxheFZhbG9yaXNcIjogMTAsIFwiUmFwcG9ydGVyaW5nXCI6IDMwfSIsIDQwLjAsICIyMDI2LTAzLTIwVDA3OjExOjEzIl0sIFsic3lubmUiLCAiU3lubmUgSMOlcnN0YWQiLCAic3lubmVAa3ZlcnZhLm5vIiwgMTIsIDIwMjYsIDAsICJ7XCJSZWduc2thcFwiOiA0MH0iLCA0MC4wLCAiMjAyNi0wMy0yMFQwNzoxMjozOCJdLCBbInJhZ25hIiwgIlJhZ25hIEZhbGthbmdlciIsICJyYWduYS5mYWxrYW5nZXJAa3ZlcnZhLm5vIiwgMTIsIDIwMjYsIDAsICJ7XCJQZWxhZ2lhXCI6IDEwLCBcIk51dHJpbWFyXCI6IDEwLCBcIkJlbmNobWFya1wiOiAxMCwgXCJJVkNcIjogMiwgXCJSYXBwb3J0ZXJpbmdcIjogOH0iLCA0MC4wLCAiMjAyNi0wMy0yMFQwNzoxNDozNSJdLCBbInRvcm9sYXYiLCAiVG9yIE9sYXYgQW5kZXJzZW4iLCAidG9iYUBrdmVydmEubm8iLCAxMiwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAzLTIwVDA3OjQwOjI5Il0sIFsicGlhIiwgIlBpYSBIYW1tZXIiLCAicGlhLmhhbW1lckBrdmVydmEubm8iLCAxMiwgMjAyNiwgMCwgIntcIlJlZ25za2FwXCI6IDQwfSIsIDQwLjAsICIyMDI2LTAzLTIwVDA4OjM5OjA0Il0sIFsidG9yZ2VpciIsICJUb3JnZWlyIFN2YWUiLCAidG9yZ2Vpci5zdmFlQGt2ZXJ2YS5ubyIsIDEyLCAyMDI2LCAwLCAie1wiUGVsYWdpYVwiOiAzLCBcIlNjYWxlXCI6IDcsIFwiSW5zdWxhXCI6IDE0LCBcIklWQ1wiOiAyLCBcIk55ZSBjYXNlIEludmVzdGVyaW5nZXJcIjogMiwgXCJTdHlyZW3DuHRlclwiOiAxNSwgXCJSZWduc2thcFwiOiAzLCBcIlJhcHBvcnRlcmluZ1wiOiAzLCBcIkt2ZXJ2YS1tw7h0ZXJcIjogM30iLCA1Mi4wLCAiMjAyNi0wMy0yMFQwODo0NToxNCJdLCBbImFtdW5kIiwgIkFtdW5kIEdhbGFlbiIsICJhbXVuZC5nYWxhZW5Aa3ZlcnZhLm5vIiwgMTIsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogMjAsIFwiU2lua2FiZXJnLUhhbnNlblwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMy0yMFQwOToxMTo1NCJdLCBbImJvcmdlIiwgIkLDuHJnZSBLbHVuZ2VyYm8iLCAiYmtAa3ZlcnZhLm5vIiwgMTIsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogMTAsIFwiTnllIGNhc2UgSW52ZXN0ZXJpbmdlclwiOiAxMCwgXCJTdHlyZW3DuHRlclwiOiAyMH0iLCA0MC4wLCAiMjAyNi0wMy0yMFQxMDozMzoyMSJdLCBbIm5pa29sYWkiLCAiTmlrb2xhaSBKZW5zZW4iLCAibmlrb2xhaS5qZW5zZW5Aa3ZlcnZhLm5vIiwgMTIsIDIwMjYsIDAsICJ7XCJJbnN1bGFcIjogMzUsIFwiUmFwcG9ydGVyaW5nXCI6IDV9IiwgNDAuMCwgIjIwMjYtMDMtMjBUMTI6NDE6MjMiXSwgWyJzaW1lbiIsICJTaW1lbiBOaWxzZW4iLCAic29uQGt2ZXJ2YS5ubyIsIDEyLCAyMDI2LCAwLCAie1wiQkVXaVwiOiAzLCBcIlNhbHZlc2VuICYgVGhhbXNcIjogMiwgXCJOeWUgY2FzZSBJbnZlc3RlcmluZ2VyXCI6IDI5LCBcIlJhcHBvcnRlcmluZ1wiOiA0LCBcIkVuZHVyXCI6IDJ9IiwgNDAuMCwgIjIwMjYtMDMtMjNUMDM6NTY6MTEiXSwgWyJlcmlrIiwgIkVyaWsgRWhyZW5wb2hsIFNhbmQiLCAiZXJpay5laHJlbnBvaGwuc2FuZEBrdmVydmEubm8iLCAxMiwgMjAyNiwgMCwgIntcIk5vcmRpc2tlIGFrc2plclwiOiAxOSwgXCJTcGFyZWJhbmtlblwiOiA2LCBcIkdsb2JhbGUgYWtzamVyXCI6IDYsIFwiU3R5cmVtw7h0ZXJcIjogMiwgXCJSYXBwb3J0ZXJpbmdcIjogNSwgXCJLdmVydmEtbcO4dGVyXCI6IDJ9IiwgNDAuMCwgIjIwMjYtMDMtMjNUMDg6MzA6MDgiXSwgWyJtYXRzLm0iLCAiTWF0cyBNYWx2aWciLCAibWF0cy5tYWx2aWdAa3ZlcnZhLm5vIiwgMTIsIDIwMjYsIDAsICJ7XCJTYWxNYXJcIjogNSwgXCJJVkNcIjogMzV9IiwgNDAuMCwgIjIwMjYtMDMtMjVUMTA6NTU6MTEiXSwgWyJtYWdudXMiLCAiTWFnbnVzIER5YnZhZCIsICJtZEBrdmVydmEubm8iLCAxMiwgMjAyNiwgMCwgIntcIlNhbE1hclwiOiAxMCwgXCJOdXRyaW1hclwiOiA1LCBcIlN0eXJlbcO4dGVyXCI6IDE1LCBcIlJhcHBvcnRlcmluZ1wiOiA1LCBcIkt2ZXJ2YS1tw7h0ZXJcIjogNX0iLCA0MC4wLCAiMjAyNi0wMy0yN1QwNzo1OTozMSJdXX0="
    data = _json.loads(base64.b64decode(DATA_B64).decode())
    with db() as con:
        cols = [r[1] for r in con.execute("PRAGMA table_info(brukere)").fetchall()]
        if "aktiv" not in cols:
            con.execute("ALTER TABLE brukere ADD COLUMN aktiv INTEGER NOT NULL DEFAULT 1")
        con.execute("DELETE FROM brukere")
        con.execute("DELETE FROM svar")
        for b in data["brukere"]:
            con.execute("INSERT INTO brukere (token,navn,epost,team,aktiv) VALUES (?,?,?,?,1)", b)
        for s in data["svar"]:
            con.execute("INSERT OR IGNORE INTO svar (token,navn,epost,uke,år,fravar,timer,total,tidspunkt) VALUES (?,?,?,?,?,?,?,?,?)", s)
    with db() as con:
        nb = con.execute("SELECT COUNT(*) FROM brukere").fetchone()[0]
        ns = con.execute("SELECT COUNT(*) FROM svar").fetchone()[0]
    return JSONResponse({"ok": True, "brukere": nb, "svar": ns})

@app.post("/admin/brukere/sett-aktiv")
async def admin_sett_aktiv(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    token = form.get("token", "").strip()
    aktiv = int(form.get("aktiv", "1"))
    sett_aktiv_bruker(token, aktiv)
    with db() as con:
        navn = con.execute("SELECT navn FROM brukere WHERE token=?", (token,)).fetchone()
    navn = navn["navn"] if navn else token
    status = "aktivert" if aktiv else "deaktivert"
    return RedirectResponse(f"/admin?melding={navn}+{status}", status_code=303)

@app.post("/admin/investeringer/legg-til")
async def admin_legg_til_inv(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    navn     = form.get("navn", "").strip()
    kategori = form.get("kategori", "Annet").strip()
    if navn:
        inv = les_investeringer()
        if navn not in [i["navn"] for i in inv]:
            inv.append({"navn": navn, "kategori": kategori})
            lagre_investeringer(inv)
    return RedirectResponse(f"/admin?melding={navn}+lagt+til", status_code=303)

@app.post("/admin/investeringer/fjern")
async def admin_fjern_inv(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    navn = form.get("navn", "").strip()
    lagre_investeringer([i for i in les_investeringer() if i["navn"] != navn])
    return RedirectResponse(f"/admin?melding={navn}+fjernet", status_code=303)

@app.post("/admin/brukere/sett-team")
async def admin_sett_team(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    token = form.get("token", "").strip()
    team = form.get("team", "investering").strip()
    if team not in ("investering", "støtte"):
        team = "investering"
    sett_team_bruker(token, team)
    return RedirectResponse("/admin?melding=Team+oppdatert", status_code=303)

@app.post("/admin/brukere/sett-lønn")
async def admin_sett_lønn(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    token = form.get("token", "").strip()
    raw = form.get("lønn", "0").replace(" ", "").replace("\u00a0", "").replace(",", "").replace(".", "") or "0"
    try:
        lønn = int(raw)
    except ValueError:
        lønn = 0
    sett_lønn_bruker(token, lønn)
    return RedirectResponse("/admin?melding=Lønn+oppdatert", status_code=303)

@app.post("/admin/investeringer/reorder")
async def admin_reorder_inv(request: Request):
    if not er_innlogget(request):
        return JSONResponse({"ok": False}, status_code=401)
    data = await request.json()
    lagre_investeringer(data)
    return JSONResponse({"ok": True})

@app.post("/admin/investeringer/endre-kategori")
async def admin_endre_kategori(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    navn     = form.get("navn", "").strip()
    kategori = form.get("kategori", "Annet").strip()
    inv = les_investeringer()
    for i in inv:
        if i["navn"] == navn:
            i["kategori"] = kategori
            break
    lagre_investeringer(inv)
    return RedirectResponse(f"/admin?melding={navn}+oppdatert", status_code=303)


@app.get("/api/brukere")
async def api_brukere(request: Request, key: Optional[str] = Query(default=None)):
    api_key_ok = EXPORT_API_KEY and key and secrets.compare_digest(key, EXPORT_API_KEY)
    if not api_key_ok:
        return JSONResponse({"error": "Ikke autorisert"}, status_code=403)
    base = str(request.base_url).rstrip("/")
    with db() as con:
        rader = con.execute("SELECT token, navn, epost FROM brukere ORDER BY navn").fetchall()
    return JSONResponse([{
        "token": r["token"],
        "navn":  r["navn"],
        "epost": r["epost"],
        "link":  f"{base}/puls/{r['token']}",
    } for r in rader])

@app.get("/admin/fordeling", response_class=HTMLResponse)
async def admin_fordeling_get(request: Request,
                               total_kostnad: Optional[float] = Query(None),
                               måned: Optional[int] = Query(None),
                               år: Optional[int] = Query(None)):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    beregning = None
    feil = None
    if total_kostnad is not None and måned is not None and år is not None:
        try:
            beregning = beregn_fordeling(total_kostnad, måned, år)
        except Exception as e:
            feil = str(e)
    today = date.today()
    return render("fordeling.html",
        beregning=beregning,
        feil=feil,
        default_måned=måned or today.month,
        default_år=år or today.year,
        default_kostnad=total_kostnad or "",
        måneder=list(enumerate(MÅNEDS_NAVN, 1)),
    )

@app.get("/admin/fordeling/eksport.csv")
async def admin_fordeling_eksport(request: Request,
                                   total_kostnad: float = Query(...),
                                   måned: int = Query(...),
                                   år: int = Query(...)):
    if not er_innlogget(request):
        return HTMLResponse("Ikke innlogget", status_code=403)
    beregning = beregn_fordeling(total_kostnad, måned, år)
    dato_str = date(år, måned, 1).strftime("%Y-%m-%d")
    linjer = ["investering,dato,sum,kommentar"]
    for rad in beregning["resultat"]:
        kommentar = f"{rad['prosent']}% av total"
        linjer.append(f'"{rad["investering"]}",{dato_str},{rad["kostnad"]},"{kommentar}"')
    csv_data = "\n".join(linjer)
    return Response(
        content=csv_data.encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename=fordeling_{år}-{måned:02d}.csv'},
    )

@app.get("/admin/eksport.csv")
async def eksport_csv(request: Request, key: Optional[str] = Query(default=None)):
    api_key_ok = EXPORT_API_KEY and key and secrets.compare_digest(key, EXPORT_API_KEY)
    if not api_key_ok and not er_innlogget(request):
        return HTMLResponse("Ikke innlogget", status_code=403)
    with db() as con:
        rader = con.execute("SELECT * FROM svar ORDER BY år, uke, navn").fetchall()
    linjer = ["Navn,Epost,Uke,År,Investering,Timer,Tidspunkt"]
    for r in rader:
        if r["fravar"]:
            linjer.append(f'{r["navn"]},{r["epost"]},{r["uke"]},{r["år"]},Fravær,0,{r["tidspunkt"]}')
        else:
            timer = json.loads(r["timer"])
            for inv, t in timer.items():
                linjer.append(f'{r["navn"]},{r["epost"]},{r["uke"]},{r["år"]},{inv},{t},{r["tidspunkt"]}')
    csv_data = "\n".join(linjer)
    return Response(
        content=csv_data.encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=puls_eksport.csv"},
    )


# ── Trivsel ───────────────────────────────────────────────────────────────────

TRIVSEL_MIN_SVAR = 3

def trivsel_er_stengt(u) -> bool:
    if u["stengt"]:
        return True
    try:
        åpnet = datetime.fromisoformat(u["opprettet"])
        return datetime.now() > åpnet + timedelta(days=u["åpen_dager"])
    except Exception:
        return False

def trivsel_opprett_utsendelse(år: int, måned: int) -> tuple[int, list]:
    """Oppretter utsendelse + survey_tokens for alle aktive brukere. Idempotent."""
    with db() as con:
        rad = con.execute(
            "SELECT id FROM trivsel_utsendelser WHERE år=? AND måned=?", (år, måned)
        ).fetchone()
        if rad:
            uid = rad["id"]
        else:
            con.execute(
                "INSERT INTO trivsel_utsendelser (måned, år, opprettet) VALUES (?,?,?)",
                (måned, år, datetime.now().isoformat())
            )
            uid = con.execute("SELECT last_insert_rowid()").fetchone()[0]

        brukere_rader = con.execute("SELECT token, navn, epost FROM brukere WHERE aktiv=1").fetchall()
        resultat = []
        for b in brukere_rader:
            eks = con.execute(
                "SELECT survey_token FROM trivsel_tokens WHERE utsendelse_id=? AND bruker_token=?",
                (uid, b["token"])
            ).fetchone()
            if eks:
                stok = eks["survey_token"]
            else:
                stok = secrets.token_urlsafe(24)
                con.execute(
                    "INSERT INTO trivsel_tokens (survey_token, utsendelse_id, bruker_token) VALUES (?,?,?)",
                    (stok, uid, b["token"])
                )
            resultat.append({"navn": b["navn"], "epost": b["epost"], "survey_token": stok})
    return uid, resultat

def trivsel_hent_resultater(utsendelse_id: int) -> dict:
    with db() as con:
        rader = con.execute(
            "SELECT trivsel, samarbeid FROM trivsel_svar WHERE utsendelse_id=?",
            (utsendelse_id,)
        ).fetchall()
    antall = len(rader)
    if antall < TRIVSEL_MIN_SVAR:
        return {"antall": antall, "skjult": True}
    snitt_t = sum(r["trivsel"]   for r in rader) / antall
    snitt_s = sum(r["samarbeid"] for r in rader) / antall
    dist_t = {i: 0 for i in range(1, 8)}
    dist_s = {i: 0 for i in range(1, 8)}
    for r in rader:
        dist_t[r["trivsel"]]   += 1
        dist_s[r["samarbeid"]] += 1
    return {
        "antall": antall,
        "skjult": False,
        "snitt_trivsel":   round(snitt_t, 1),
        "snitt_samarbeid": round(snitt_s, 1),
        "dist_trivsel":    dist_t,
        "dist_samarbeid":  dist_s,
    }

@app.get("/trivsel/takk", response_class=HTMLResponse)
async def trivsel_takk_get():
    return render("trivsel_takk.html")

@app.get("/trivsel/allerede-svart", response_class=HTMLResponse)
async def trivsel_allerede_svart_get():
    return render("trivsel_allerede_svart.html", måned="", år="")

@app.get("/trivsel/{survey_token}", response_class=HTMLResponse)
async def trivsel_vis_skjema(survey_token: str):
    with db() as con:
        row = con.execute("""
            SELECT tt.id, tt.brukt, tt.utsendelse_id,
                   tu.måned, tu.år, tu.stengt, tu.åpen_dager, tu.opprettet
            FROM trivsel_tokens tt
            JOIN trivsel_utsendelser tu ON tu.id = tt.utsendelse_id
            WHERE tt.survey_token = ?
        """, (survey_token,)).fetchone()
    if not row:
        return render("trivsel_feil.html", melding="Lenken er ugyldig eller utløpt.")
    måned_navn = MÅNEDS_NAVN[row["måned"] - 1]
    if trivsel_er_stengt(row):
        return render("trivsel_stengt.html", måned=måned_navn, år=row["år"])
    if row["brukt"]:
        return render("trivsel_allerede_svart.html", måned=måned_navn, år=row["år"])
    return render("trivsel_svar.html",
                  survey_token=survey_token,
                  måned=måned_navn, år=row["år"],
                  forhåndsvis=False)

@app.post("/trivsel/{survey_token}")
async def trivsel_send_svar(survey_token: str, request: Request):
    form = await request.form()
    try:
        trivsel_score   = int(form.get("trivsel",   0))
        samarbeid_score = int(form.get("samarbeid", 0))
    except (TypeError, ValueError):
        from fastapi import HTTPException as _H
        raise _H(400, "Ugyldig input")
    if not (1 <= trivsel_score <= 7 and 1 <= samarbeid_score <= 7):
        from fastapi import HTTPException as _H
        raise _H(400, "Score må være mellom 1 og 7")
    with db() as con:
        tok = con.execute(
            "SELECT id, brukt, utsendelse_id FROM trivsel_tokens WHERE survey_token=?",
            (survey_token,)
        ).fetchone()
        if not tok:
            from fastapi import HTTPException as _H
            raise _H(404)
        if tok["brukt"]:
            return RedirectResponse("/trivsel/allerede-svart", status_code=303)
        con.execute("UPDATE trivsel_tokens SET brukt=1 WHERE survey_token=?", (survey_token,))
        con.execute(
            "INSERT INTO trivsel_svar (utsendelse_id, trivsel, samarbeid, innsendt) VALUES (?,?,?,?)",
            (tok["utsendelse_id"], trivsel_score, samarbeid_score, datetime.now().isoformat())
        )
    return RedirectResponse("/trivsel/takk", status_code=303)

@app.get("/admin/trivsel", response_class=HTMLResponse)
async def admin_trivsel(request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    nå = datetime.now()

    # Auto-opprett alle måneder Jan–nåværende for inneværende år
    for m in range(1, nå.month + 1):
        trivsel_opprett_utsendelse(nå.year, m)
        # Auto-steng måneder eldre enn 10 dager etter siste dag
        with db() as con:
            u = con.execute(
                "SELECT id, stengt, opprettet, åpen_dager FROM trivsel_utsendelser WHERE år=? AND måned=?",
                (nå.year, m)
            ).fetchone()
            if u and not u["stengt"] and m < nå.month:
                # Steng foregående måneder automatisk hvis de er eldre
                try:
                    åpnet = datetime.fromisoformat(u["opprettet"])
                    if nå > åpnet + timedelta(days=u["åpen_dager"]):
                        con.execute("UPDATE trivsel_utsendelser SET stengt=1 WHERE id=?", (u["id"],))
                except Exception:
                    pass

    with db() as con:
        utsendelser = con.execute(
            "SELECT * FROM trivsel_utsendelser ORDER BY år DESC, måned DESC"
        ).fetchall()
        antall_brukere = con.execute("SELECT COUNT(*) FROM brukere").fetchone()[0]

    måneder_data = []
    for u in utsendelser:
        res = trivsel_hent_resultater(u["id"])
        with db() as con:
            totalt = con.execute("SELECT COUNT(*) FROM trivsel_tokens WHERE utsendelse_id=?", (u["id"],)).fetchone()[0]
            svart  = con.execute("SELECT COUNT(*) FROM trivsel_tokens WHERE utsendelse_id=? AND brukt=1", (u["id"],)).fetchone()[0]
        stengt_nå = trivsel_er_stengt(u)
        måneder_data.append({
            "id":       u["id"],
            "måned":    MÅNEDS_NAVN[u["måned"] - 1],
            "måned_nr": u["måned"],
            "år":       u["år"],
            "totalt":   totalt,
            "svart":    svart,
            "pst":      round(svart / totalt * 100) if totalt else 0,
            "res":      res,
            "stengt":   stengt_nå,
        })

    return render("admin_trivsel.html",
                  måneder=måneder_data,
                  antall_brukere=antall_brukere,
                  default_måned=nå.month,
                  default_år=nå.year,
                  månednavn=MÅNEDS_NAVN,
                  min_svar=TRIVSEL_MIN_SVAR,
                  melding=request.query_params.get("melding", ""))

@app.post("/admin/trivsel/start")
async def admin_trivsel_start(request: Request):
    """Start eller gjenåpne en trivsel-runde for valgt måned/år."""
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    try:
        år   = int(form.get("år",   0))
        måned = int(form.get("måned", 0))
    except (ValueError, TypeError):
        return RedirectResponse("/admin/trivsel?melding=Ugyldig+dato", status_code=303)
    if not (1 <= måned <= 12 and 2020 <= år <= 2035):
        return RedirectResponse("/admin/trivsel?melding=Ugyldig+dato", status_code=303)
    trivsel_opprett_utsendelse(år, måned)
    return RedirectResponse(f"/admin/trivsel/lenker/{år}/{måned}", status_code=303)

@app.post("/admin/trivsel/steng/{uid}")
async def admin_trivsel_steng(uid: int, request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    with db() as con:
        con.execute("UPDATE trivsel_utsendelser SET stengt=1 WHERE id=?", (uid,))
    return RedirectResponse("/admin/trivsel?melding=Periode+stengt", status_code=303)

@app.get("/admin/trivsel/forhåndsvis", response_class=HTMLResponse)
async def admin_trivsel_preview(request: Request):
    """Vis skjemaet slik brukerne ser det — uten å registrere svar."""
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    nå = datetime.now()
    return render("trivsel_svar.html",
                  survey_token="__PREVIEW__",
                  måned=MÅNEDS_NAVN[nå.month - 1],
                  år=nå.year,
                  forhåndsvis=True)

@app.post("/admin/trivsel/testdata/{uid}")
async def admin_trivsel_testdata(uid: int, request: Request):
    """Generer 5 tilfeldige testsvar for en utsendelse (vises som demo-data)."""
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    import random as _r
    with db() as con:
        u = con.execute("SELECT id, måned, år FROM trivsel_utsendelser WHERE id=?", (uid,)).fetchone()
        if not u:
            return RedirectResponse("/admin/trivsel", status_code=303)
        # Hent tokens som ikke er brukt
        ubrukte = con.execute(
            "SELECT survey_token, bruker_token FROM trivsel_tokens WHERE utsendelse_id=? AND brukt=0",
            (uid,)
        ).fetchall()
        antall = min(5, len(ubrukte))
        valgte = _r.sample(list(ubrukte), antall)
        for t in valgte:
            trivsel_score   = _r.randint(4, 7)
            samarbeid_score = _r.randint(4, 7)
            con.execute("UPDATE trivsel_tokens SET brukt=1 WHERE survey_token=?", (t["survey_token"],))
            con.execute(
                "INSERT INTO trivsel_svar (utsendelse_id, trivsel, samarbeid, innsendt) VALUES (?,?,?,?)",
                (uid, trivsel_score, samarbeid_score, datetime.now().isoformat())
            )
    return RedirectResponse(f"/admin/trivsel?melding={antall}+testsvar+lagt+inn", status_code=303)

@app.post("/admin/trivsel/nullstill-svar")
async def admin_trivsel_nullstill(request: Request):
    """Nullstill én persons svar — for testing."""
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    survey_token = (form.get("survey_token") or "").strip()
    year  = int(form.get("year",  0) or 0)
    month = int(form.get("month", 0) or 0)
    if survey_token:
        with db() as con:
            tok = con.execute(
                "SELECT id, utsendelse_id FROM trivsel_tokens WHERE survey_token=? AND brukt=1",
                (survey_token,)
            ).fetchone()
            if tok:
                con.execute("""
                    DELETE FROM trivsel_svar WHERE id = (
                        SELECT id FROM trivsel_svar
                        WHERE utsendelse_id=? ORDER BY id DESC LIMIT 1
                    )
                """, (tok["utsendelse_id"],))
                con.execute("UPDATE trivsel_tokens SET brukt=0 WHERE id=?", (tok["id"],))
    return RedirectResponse(f"/admin/trivsel/lenker/{year}/{month}", status_code=303)

@app.get("/admin/trivsel/lenker/{year}/{month}", response_class=HTMLResponse)
async def admin_trivsel_lenker(year: int, month: int, request: Request):
    if not er_innlogget(request):
        return RedirectResponse("/admin", status_code=303)
    trivsel_opprett_utsendelse(year, month)  # idempotent — sikrer tokens finnes
    with db() as con:
        u = con.execute(
            "SELECT id FROM trivsel_utsendelser WHERE år=? AND måned=?", (year, month)
        ).fetchone()
        rader = con.execute("""
            SELECT b.navn, b.epost, tt.survey_token, tt.brukt
            FROM trivsel_tokens tt
            JOIN brukere b ON b.token = tt.bruker_token
            WHERE tt.utsendelse_id = ?
            ORDER BY tt.brukt ASC, b.navn
        """, (u["id"],)).fetchall()
    base = str(request.base_url).rstrip("/")
    lenker = [
        {
            "navn":         r["navn"],
            "epost":        r["epost"],
            "link":         f"{base}/trivsel/{r['survey_token']}",
            "survey_token": r["survey_token"],
            "brukt":        bool(r["brukt"]),
        }
        for r in rader
    ]
    svart = sum(1 for l in lenker if l["brukt"])
    return render("trivsel_lenker.html",
                  lenker=lenker,
                  måned=MÅNEDS_NAVN[month - 1],
                  måned_nr=month,
                  år=year,
                  svart=svart)

@app.get("/api/trivsel/lenker/{year}/{month}")
async def api_trivsel_lenker(year: int, month: int, request: Request):
    """Power Automate: hent survey-lenker for utsending via e-post."""
    key = request.query_params.get("api_key") or request.headers.get("x-api-key", "")
    if not (EXPORT_API_KEY and secrets.compare_digest(key, EXPORT_API_KEY)):
        from fastapi import HTTPException as _H
        raise _H(401, "Ugyldig API-nøkkel")
    _, tokens = trivsel_opprett_utsendelse(year, month)
    base = str(request.base_url).rstrip("/")
    return JSONResponse([
        {"navn": t["navn"], "epost": t["epost"], "link": f"{base}/trivsel/{t['survey_token']}"}
        for t in tokens
    ])


if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8502, reload=True, app_dir=str(BASE))
